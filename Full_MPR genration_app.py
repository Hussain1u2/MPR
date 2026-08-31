"""
ChargeZone MPR - Issue Tracker Dashboard
Executive Streamlit Application for Operational Issue Tracking, SLA Governance & Field Analytics.
"""
import calendar
from datetime import datetime
from io import BytesIO
import os
import re

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st

st.set_page_config(
    page_title="ChargeZone | MPR - Issue Tracker Dashboard",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─── Constants & Configurations ───────────────────────────────────────────────

BAD_VALUES = {
    '', '#REF!', '#N/A', '#VALUE!', '#NAME?', 'NONE', 'NULL', 'NAN', 'NAT',
    'N/A', 'NA', 'N.A.', 'N.A', '<NA>', '#N/A N/A', 'UNKNOWN', 'UNDEFINED',
    'REF!', '#NULL!', '#NUM!', '#DIV/0!', '-', '--', 'NONE', '0'
}

_MONTH_NUM = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

# Modern executive color palette
COLOR_PRIMARY = "#DC2626"       # Crimson Red
COLOR_DARK = "#991B1B"          # Deep Burgundy
COLOR_ACCENT = "#7F1D1D"        # Dark Cherry
COLOR_SUCCESS = "#10B981"       # Emerald Green
COLOR_WARNING = "#F59E0B"       # Amber Orange
COLOR_DANGER = "#EF4444"        # Red Alert
COLOR_INFO = "#3B82F6"          # Royal Blue
COLOR_INDIGO = "#6366F1"        # Indigo
COLOR_PURPLE = "#8B5CF6"        # Purple
COLOR_SLATE = "#64748B"         # Slate Grey
COLOR_LIGHT_BG = "#F8FAFC"      # Clean Light Background
COLOR_CARD_BG = "#FFFFFF"

PALETTE_VIBRANT = ["#DC2626", "#2563EB", "#10B981", "#F59E0B", "#8B5CF6", "#EC4899", "#06B6D4", "#64748B"]
PALETTE_STATUS = {"Open": "#EF4444", "In Progress": "#F59E0B", "Resolved": "#10B981", "Closed": "#1E293B"}
PALETTE_SLA = {"Within TAT": "#10B981", "Breached TAT": "#EF4444", "Pending": "#F59E0B"}

# ─── Helper Functions ─────────────────────────────────────────────────────────

def clean_val(v):
    if v is None or pd.isna(v):
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.upper() in BAD_VALUES:
            return None
        return s
    return v

def norm_header(h):
    if h is None:
        return ''
    return ''.join(c for c in str(h).lower() if c.isalnum())

def find_col(df, possible_names):
    if df is None or df.empty:
        return None
    clean_cols = {norm_header(c): c for c in df.columns}
    for p in possible_names:
        p_norm = norm_header(p)
        if p_norm in clean_cols:
            return clean_cols[p_norm]
    for p in possible_names:
        p_norm = norm_header(p)
        for c_norm, orig in clean_cols.items():
            if p_norm in c_norm or c_norm in p_norm:
                return orig
    return None

def parse_date_val(val):
    """Robust date parser handling Timestamp, datetime, Excel serial integers, and date strings."""
    if val is None or pd.isna(val):
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.Timestamp(val)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        num_val = float(val)
        if 35000 <= num_val <= 70000:
            try:
                dt_conv = pd.to_datetime(num_val, unit='D', origin='1899-12-30')
                if not pd.isna(dt_conv):
                    return pd.Timestamp(dt_conv)
            except Exception:
                pass
    s = str(val).strip()
    if not s or s.upper() in BAD_VALUES:
        return None
    if re.match(r'^\d{4}[-/]', s):
        parsed_dt = pd.to_datetime(s, errors='coerce', dayfirst=False)
    else:
        parsed_dt = pd.to_datetime(s, errors='coerce', dayfirst=True)
    return None if pd.isna(parsed_dt) else pd.Timestamp(parsed_dt)

# ─── Issue Tracker Excel Parser ───────────────────────────────────────────────

def select_sheet_name(sheetnames, preferred_names, fallback_keyword):
    if not sheetnames:
        return None
    clean_map = {str(s).strip().lower(): s for s in sheetnames if s is not None}
    for pref in preferred_names:
        pref_clean = pref.strip().lower()
        if pref_clean in clean_map:
            return clean_map[pref_clean]
    for pref in preferred_names:
        pref_clean = pref.strip().lower()
        for s in sheetnames:
            if s and pref_clean in str(s).strip().lower():
                return s
    for s in sheetnames:
        if s and fallback_keyword.lower() in str(s).strip().lower():
            return s
    return sheetnames[0]


def parse_issue_tracker(wb_or_bytes):
    """Extract, clean, and enrich all fields from the Issue Tracker worksheet."""
    if isinstance(wb_or_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(wb_or_bytes), read_only=True, data_only=True)
        should_close = True
    else:
        wb = wb_or_bytes
        should_close = False

    try:
        sheet_name = select_sheet_name(wb.sheetnames, ['Issue Tracker', 'Issue Data', 'Issues', 'Issue_Tracker'], 'issue')
        if not sheet_name or sheet_name not in wb.sheetnames:
            return pd.DataFrame()

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()

        header_idx = 0
        for idx, r in enumerate(rows[:15]):
            row_norms = [norm_header(c) for c in r if c is not None]
            if 'issueid' in row_norms or 'srno' in row_norms or ('zone' in row_norms and 'status' in row_norms):
                header_idx = idx
                break

        header_row = rows[header_idx]
        idx_map = {norm_header(h): i for i, h in enumerate(header_row) if h is not None}

        # Key Index Mappings
        col_srno = idx_map.get('srno', idx_map.get('srno.'))
        col_issueid = idx_map.get('issueid', idx_map.get('ticketid', idx_map.get('srno')))
        col_ocpp = idx_map.get('ocppid', idx_map.get('chargerid'))
        col_jira = idx_map.get('jirafdticket', idx_map.get('jiraticket', idx_map.get('fdticket')))
        col_severity = idx_map.get('severity', idx_map.get('priority'))
        col_issuedate = idx_map.get('issuedate', idx_map.get('createddate', idx_map.get('date')))
        col_status = idx_map.get('status', idx_map.get('ticketstatus', idx_map.get('issuestatus')))
        col_type = idx_map.get('issuetype', idx_map.get('category', idx_map.get('faulttype')))
        col_subtype = idx_map.get('issuesubtype', idx_map.get('subtype', idx_map.get('faultsubtype')))
        col_desc = idx_map.get('issuedescription', idx_map.get('description', idx_map.get('faultdetails')))
        col_resdate = idx_map.get('resolutiondate', idx_map.get('resolveddate', idx_map.get('closeddate')))
        col_restdate = idx_map.get('restorationdate', idx_map.get('restoreddate'))
        col_action = idx_map.get('correctiveactiontaken', idx_map.get('correctiveaction', idx_map.get('actiontaken')))
        col_remarks = idx_map.get('remarks1', idx_map.get('remarks', idx_map.get('remark')))
        col_state = idx_map.get('state', idx_map.get('province'))
        col_zone = idx_map.get('zone', idx_map.get('region'))
        col_segment = idx_map.get('b2bb2c', idx_map.get('segment', idx_map.get('customersegment')))
        col_zme = idx_map.get('zme', idx_map.get('leadzme', idx_map.get('managername', idx_map.get('zonemanager'))))
        col_manager = idx_map.get('managername', idx_map.get('reportingmanager', idx_map.get('leadmanager')))
        col_dep = idx_map.get('dependency', idx_map.get('owner', idx_map.get('assignee')))
        col_stationid = idx_map.get('stationid', idx_map.get('siteid', idx_map.get('stationcode')))
        col_stationname = idx_map.get('stationname', idx_map.get('sitename', idx_map.get('location')))
        col_tatdays = idx_map.get('tatdays', idx_map.get('sladays', idx_map.get('targettat')))
        col_tdoc = idx_map.get('tdoc', idx_map.get('targetdate', idx_map.get('duedate')))
        col_report = idx_map.get('reportfilling', idx_map.get('reportrequired'))
        col_age = idx_map.get('ageofissue', idx_map.get('age', idx_map.get('ticketage')))
        col_tat = idx_map.get('tatcompliance', idx_map.get('slacompliance', idx_map.get('compliance')))
        col_make = idx_map.get('chargermake', idx_map.get('make', idx_map.get('oem', idx_map.get('chargeroem'))))

        records = []
        for r_idx, r in enumerate(rows[header_idx + 1:], start=1):
            if not r or all(c is None for c in r):
                continue

            raw_issue_id = r[col_issueid] if col_issueid is not None and col_issueid < len(r) else None
            issue_id = clean_val(raw_issue_id)
            if issue_id is None:
                raw_ocpp = r[col_ocpp] if col_ocpp is not None and col_ocpp < len(r) else None
                if clean_val(raw_ocpp) is not None:
                    issue_id = f"CZ-{r_idx:04d}"
                else:
                    continue

            # Date Parsing
            raw_issue_date = r[col_issuedate] if col_issuedate is not None and col_issuedate < len(r) else None
            issue_date = parse_date_val(raw_issue_date)

            raw_res_date = r[col_resdate] if col_resdate is not None and col_resdate < len(r) else None
            resolution_date = parse_date_val(raw_res_date)

            raw_rest_date = r[col_restdate] if col_restdate is not None and col_restdate < len(r) else None
            restoration_date = parse_date_val(raw_rest_date)

            raw_tdoc = r[col_tdoc] if col_tdoc is not None and col_tdoc < len(r) else None
            tdoc_date = parse_date_val(raw_tdoc)

            # Month and Week formatting
            if issue_date is not None:
                mlabel = issue_date.strftime('%b-%y')
                iso_year, iso_week, _ = issue_date.isocalendar()
                start_dt = pd.Timestamp(datetime.fromisocalendar(iso_year, iso_week, 1)).to_pydatetime()
                end_dt = pd.Timestamp(datetime.fromisocalendar(iso_year, iso_week, 7)).to_pydatetime()
                week_range = f"W{iso_week:02d} ({start_dt.strftime('%d-%b')} to {end_dt.strftime('%d-%b-%Y')})"
                schedule_week = f"W{iso_week:02d}"
            else:
                mlabel = 'Unscheduled'
                week_range = "Unscheduled"
                schedule_week = "Unscheduled"
                start_dt = None
                end_dt = None

            # Status & Severity
            status_raw = str(clean_val(r[col_status]) if col_status is not None and col_status < len(r) else 'Open').strip()
            if status_raw.upper() in ['RESOLVED', 'CLOSED', 'COMPLETE', 'COMPLETED']:
                norm_status = 'Closed' if 'CLOSE' in status_raw.upper() else 'Resolved'
            elif status_raw.upper() in ['OPEN', 'PENDING', 'NEW', 'ACTIVE']:
                norm_status = 'Open'
            elif 'PROGRESS' in status_raw.upper() or 'HOLD' in status_raw.upper():
                norm_status = 'In Progress'
            else:
                norm_status = status_raw.title()

            severity_raw = str(clean_val(r[col_severity]) if col_severity is not None and col_severity < len(r) else 'Minor').strip().title()
            if severity_raw not in ['Critical', 'Major', 'Minor']:
                if 'CRIT' in severity_raw.upper():
                    severity_raw = 'Critical'
                elif 'MAJ' in severity_raw.upper():
                    severity_raw = 'Major'
                else:
                    severity_raw = 'Minor'

            # TAT Compliance
            tat_clean = clean_val(r[col_tat]) if col_tat is not None and col_tat < len(r) else None
            tat_raw = str(tat_clean).strip().upper() if tat_clean is not None else ''

            # Numerical TAT Days & Age
            raw_tat_days = r[col_tatdays] if col_tatdays is not None and col_tatdays < len(r) else None
            try:
                tat_days = float(raw_tat_days) if raw_tat_days is not None and not pd.isna(raw_tat_days) else None
            except Exception:
                tat_days = None

            raw_age = r[col_age] if col_age is not None and col_age < len(r) else None
            try:
                age_days = float(raw_age) if raw_age is not None and not pd.isna(raw_age) else None
                if age_days is not None and age_days < 0:
                    age_days = abs(age_days)
            except Exception:
                age_days = None

            # Resolution calculation if dates exist
            res_duration = None
            if issue_date is not None and resolution_date is not None:
                diff_days = (resolution_date - issue_date).total_seconds() / 86400.0
                res_duration = max(0.0, round(diff_days, 1))

            records.append({
                'issueId': str(issue_id),
                'ocppId': str(clean_val(r[col_ocpp]) if col_ocpp is not None and col_ocpp < len(r) else 'N/A'),
                'jiraTicket': str(clean_val(r[col_jira]) if col_jira is not None and col_jira < len(r) else '—'),
                'severity': severity_raw,
                'issueDate': issue_date,
                'status': norm_status,
                'issueType': str(clean_val(r[col_type]) if col_type is not None and col_type < len(r) else 'General').replace('_', ' ').strip(),
                'issueSubType': str(clean_val(r[col_subtype]) if col_subtype is not None and col_subtype < len(r) else 'General').replace('_', ' ').strip(),
                'description': str(clean_val(r[col_desc]) if col_desc is not None and col_desc < len(r) else 'No description provided'),
                'resolutionDate': resolution_date,
                'restorationDate': restoration_date,
                'correctiveAction': str(clean_val(r[col_action]) if col_action is not None and col_action < len(r) else '—'),
                'remarks': str(clean_val(r[col_remarks]) if col_remarks is not None and col_remarks < len(r) else '—'),
                'state': str(clean_val(r[col_state]) if col_state is not None and col_state < len(r) else 'Unknown'),
                'zone': str(clean_val(r[col_zone]) if col_zone is not None and col_zone < len(r) else 'Unknown'),
                'segment': str(clean_val(r[col_segment]) if col_segment is not None and col_segment < len(r) else 'B2C'),
                'zme': str(clean_val(r[col_zme]) if col_zme is not None and col_zme < len(r) else 'Unassigned'),
                'managerName': str(clean_val(r[col_manager]) if col_manager is not None and col_manager < len(r) else 'Unassigned'),
                'dependency': str(clean_val(r[col_dep]) if col_dep is not None and col_dep < len(r) else 'Internal'),
                'stationId': str(clean_val(r[col_stationid]) if col_stationid is not None and col_stationid < len(r) else 'N/A'),
                'stationName': str(clean_val(r[col_stationname]) if col_stationname is not None and col_stationname < len(r) else 'N/A'),
                'tatDays': tat_days,
                'tdoc': tdoc_date,
                'reportFilling': str(clean_val(r[col_report]) if col_report is not None and col_report < len(r) else 'Not Required'),
                'ageOfIssue': age_days,
                'month': mlabel,
                'scheduleWeek': schedule_week,
                'weekStartDate': start_dt,
                'weekEndDate': end_dt,
                'scheduleWeekRange': week_range,
                'tatCompliance': tat_raw,
                'chargerMake': str(clean_val(r[col_make]) if col_make is not None and col_make < len(r) else 'Unknown').strip(),
                'resolutionDurationDays': res_duration,
            })

        df = pd.DataFrame(records)
        if df.empty:
            return df

        # Normalized derived Boolean flags
        status_upper = df['status'].astype(str).str.strip().str.upper()
        df['_Is_Closed_'] = status_upper.isin(['CLOSED', 'RESOLVED', 'COMPLETE'])
        df['_Is_Open_'] = ~df['_Is_Closed_']

        tat_upper = df['tatCompliance'].astype(str).str.strip().str.upper()
        df['_Is_Within_'] = (tat_upper == 'YES') | ((df['_Is_Closed_']) & (df['resolutionDurationDays'].notna()) & (df['tatDays'].notna()) & (df['resolutionDurationDays'] <= df['tatDays']))
        df['_Is_Without_'] = (tat_upper == 'NO') | ((df['_Is_Closed_']) & (df['resolutionDurationDays'].notna()) & (df['tatDays'].notna()) & (df['resolutionDurationDays'] > df['tatDays']))

        # Adjust for unresolved tickets
        df['_Is_Closed_Within_'] = df['_Is_Closed_'] & df['_Is_Within_']
        df['_Is_Closed_Without_'] = df['_Is_Closed_'] & (~df['_Is_Within_'])

        # Overdue Open Tickets
        now_ts = pd.Timestamp.now()
        df['_Is_Overdue_'] = df['_Is_Open_'] & (
            (df['tdoc'].notna() & (df['tdoc'] < now_ts)) |
            (df['tatDays'].notna() & df['ageOfIssue'].notna() & (df['ageOfIssue'] > df['tatDays']))
        )

        return df

    finally:
        if should_close:
            wb.close()


@st.cache_data(show_spinner=False)
def load_issue_data(uploaded_bytes=None):
    """Load Issue Tracker strictly from user uploaded bytes."""
    if uploaded_bytes is not None:
        return parse_issue_tracker(uploaded_bytes)
    return pd.DataFrame()


# ─── Aggregations & Analytical Tables ─────────────────────────────────────────

def get_available_months(issue_df):
    """Return chronologically ordered list of available months."""
    if issue_df.empty or 'month' not in issue_df.columns:
        return []

    if 'issueDate' in issue_df.columns:
        valid_df = issue_df[issue_df['issueDate'].notna()]
        if not valid_df.empty:
            sorted_months = valid_df.sort_values('issueDate')['month'].unique().tolist()
            for m in issue_df['month'].unique().tolist():
                if m and m != 'Unscheduled' and m not in sorted_months:
                    sorted_months.append(m)
            return [m for m in sorted_months if m and m != 'Unscheduled']

    return [m for m in issue_df['month'].unique().tolist() if m and m != 'Unscheduled']


def compute_zme_issue_table(issue_df, selected_months=None, selected_zones=None):
    if issue_df.empty:
        return pd.DataFrame()

    df = issue_df.copy()
    if selected_months:
        df = df[df['month'].isin(selected_months)]
    if selected_zones and 'All' not in selected_zones:
        df = df[df['zone'].isin(selected_zones)]

    if df.empty:
        return pd.DataFrame()

    agg = df.groupby(['zme', 'zone']).agg(
        total=('issueId', 'count'),
        open=('_Is_Open_', 'sum'),
        closed=('_Is_Closed_', 'sum'),
        within=('_Is_Closed_Within_', 'sum'),
        outside=('_Is_Closed_Without_', 'sum'),
        overdue=('_Is_Overdue_', 'sum'),
        critical=('severity', lambda s: (s == 'Critical').sum()),
    ).reset_index()

    agg['cm_efficiency'] = (agg['closed'] / agg['total'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
    agg['tat_efficiency'] = (agg['within'] / agg['total'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
    return agg.sort_values(by='total', ascending=False)


def compute_zone_issue_table(issue_df, selected_months=None):
    if issue_df.empty or 'zone' not in issue_df.columns:
        return pd.DataFrame()

    df = issue_df.copy()
    if selected_months:
        df = df[df['month'].isin(selected_months)]

    if df.empty:
        return pd.DataFrame()

    agg = df.groupby('zone').agg(
        total=('issueId', 'count'),
        open=('_Is_Open_', 'sum'),
        closed=('_Is_Closed_', 'sum'),
        within=('_Is_Closed_Within_', 'sum'),
        outside=('_Is_Closed_Without_', 'sum'),
        overdue=('_Is_Overdue_', 'sum'),
        zme_count=('zme', 'nunique'),
        station_count=('stationId', 'nunique')
    ).reset_index()

    agg['cm_efficiency'] = (agg['closed'] / agg['total'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
    agg['tat_efficiency'] = (agg['within'] / agg['total'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
    return agg.sort_values(by='total', ascending=False)


def compute_top_repetitive_faults(issue_df, top_overall=10, top_station=20):
    """Compute Top N Overall and Top N Station-wise Repetitive Faults."""
    if issue_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 1. Top Overall Repetitive Faults
    overall_agg = issue_df.groupby(['issueType', 'issueSubType']).agg(
        total_count=('issueId', 'count'),
        unique_stations=('stationName', lambda s: s[s != 'N/A'].nunique()),
        unique_chargers=('ocppId', lambda c: c[c != 'N/A'].nunique()),
        open_count=('_Is_Open_', 'sum'),
        closed_count=('_Is_Closed_', 'sum'),
        within_tat=('_Is_Closed_Within_', 'sum'),
        breached_tat=('_Is_Closed_Without_', 'sum'),
        critical_count=('severity', lambda s: (s == 'Critical').sum()),
        major_count=('severity', lambda s: (s == 'Major').sum())
    ).reset_index().sort_values('total_count', ascending=False)

    overall_agg['tat_efficiency'] = (overall_agg['within_tat'] / overall_agg['total_count'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
    overall_agg['fault_label'] = overall_agg['issueType'] + ' — ' + overall_agg['issueSubType']
    top10_df = overall_agg.head(top_overall)

    # 2. Top Repetitive Faults by Stations
    st_valid = issue_df[issue_df['stationName'] != 'N/A'].copy()
    if not st_valid.empty:
        station_agg = st_valid.groupby(['stationName', 'zone', 'zme', 'issueType', 'issueSubType']).agg(
            fault_count=('issueId', 'count'),
            open_count=('_Is_Open_', 'sum'),
            closed_count=('_Is_Closed_', 'sum'),
            within_tat=('_Is_Closed_Within_', 'sum'),
            breached_tat=('_Is_Closed_Without_', 'sum'),
            unique_chargers=('ocppId', lambda c: c[c != 'N/A'].nunique())
        ).reset_index().sort_values('fault_count', ascending=False)

        station_agg['tat_efficiency'] = (station_agg['within_tat'] / station_agg['fault_count'].replace(0, pd.NA) * 100).fillna(0.0).round(1)
        station_agg['station_fault_label'] = station_agg['stationName'].apply(lambda s: str(s)[:28] + '...' if len(str(s)) > 28 else str(s)) + ' | ' + station_agg['issueSubType']
        top20_station_df = station_agg.head(top_station)
    else:
        top20_station_df = pd.DataFrame()

    return top10_df, top20_station_df


# ─── Excel Multi-Tab Report Generator ─────────────────────────────────────────

@st.cache_data(show_spinner=False)
def generate_issue_excel_report(filtered_df, full_df, months_list, active_filters_info=None):
    """Generate executive multi-tab Excel workbook with embedded native Excel charts according to active filters."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Executive KPI Summary
        tot = len(filtered_df)
        open_cnt = int(filtered_df['_Is_Open_'].sum()) if not filtered_df.empty else 0
        closed_cnt = int(filtered_df['_Is_Closed_'].sum()) if not filtered_df.empty else 0
        within_tat = int(filtered_df['_Is_Closed_Within_'].sum()) if not filtered_df.empty else 0
        outside_tat = int(filtered_df['_Is_Closed_Without_'].sum()) if not filtered_df.empty else 0
        overdue_cnt = int(filtered_df['_Is_Overdue_'].sum()) if not filtered_df.empty else 0
        crit_cnt = int((filtered_df['severity'] == 'Critical').sum()) if not filtered_df.empty else 0
        maj_cnt = int((filtered_df['severity'] == 'Major').sum()) if not filtered_df.empty else 0

        cm_eff = (closed_cnt / tot * 100) if tot > 0 else 0.0
        tat_eff = (within_tat / tot * 100) if tot > 0 else 0.0

        period_str = ', '.join(months_list) if months_list else 'All Recorded Months'
        if active_filters_info and isinstance(active_filters_info, dict):
            filter_summary_str = f"Months: {period_str} | States: {active_filters_info.get('states', 'All')} | Zones: {active_filters_info.get('zones', 'All')} | Make: {active_filters_info.get('make', 'All')}"
        else:
            filter_summary_str = period_str

        kpi_summary = [
            {'Parameter': 'Active Filter Criteria / Scope', 'Metric Value': filter_summary_str},
            {'Parameter': 'Total Faults Registered', 'Metric Value': tot},
            {'Parameter': 'Active / Open Faults', 'Metric Value': open_cnt},
            {'Parameter': 'Resolved / Closed Faults', 'Metric Value': closed_cnt},
            {'Parameter': 'Closed Within TAT (SLA Compliant)', 'Metric Value': within_tat},
            {'Parameter': 'Closed Without TAT (SLA Breached)', 'Metric Value': outside_tat},
            {'Parameter': 'Active Overdue Tickets', 'Metric Value': overdue_cnt},
            {'Parameter': 'Critical Severity Faults', 'Metric Value': crit_cnt},
            {'Parameter': 'Major Severity Faults', 'Metric Value': maj_cnt},
            {'Parameter': 'Overall CM Efficiency %', 'Metric Value': f"{cm_eff:.2f}%"},
            {'Parameter': 'CM-TAT Efficiency %', 'Metric Value': f"{tat_eff:.2f}%"},
        ]
        pd.DataFrame(kpi_summary).to_excel(writer, sheet_name='Executive Summary', index=False)
        ws_exec = writer.sheets['Executive Summary']

        # SLA Breakdown Chart Reference Data on Executive Summary Sheet
        ws_exec.cell(row=14, column=1, value="SLA Status Category")
        ws_exec.cell(row=14, column=2, value="Incident Volume")
        ws_exec.cell(row=15, column=1, value="Closed Within TAT")
        ws_exec.cell(row=15, column=2, value=within_tat)
        ws_exec.cell(row=16, column=1, value="Closed Breached TAT")
        ws_exec.cell(row=16, column=2, value=outside_tat)
        ws_exec.cell(row=17, column=1, value="Active Open")
        ws_exec.cell(row=17, column=2, value=open_cnt)

        # Embedded Native Excel PieChart on Executive Summary
        pie_exec = PieChart()
        pie_exec.title = "Executive SLA Compliance Breakdown"
        pie_exec.width = 16
        pie_exec.height = 10
        data_exec_pie = Reference(ws_exec, min_col=2, min_row=14, max_row=17)
        labels_exec_pie = Reference(ws_exec, min_col=1, min_row=15, max_row=17)
        pie_exec.add_data(data_exec_pie, titles_from_data=True)
        pie_exec.set_categories(labels_exec_pie)
        ws_exec.add_chart(pie_exec, "D2")

        # Sheet 2: Repetitive Faults Analysis
        top10_rep, top20_st_rep = compute_top_repetitive_faults(filtered_df, top_overall=10, top_station=20)
        if not top10_rep.empty:
            top10_disp = top10_rep.rename(columns={
                'issueType': 'Category',
                'issueSubType': 'Fault Sub-Type',
                'total_count': 'Total Incidents',
                'unique_stations': 'Impacted Sites',
                'unique_chargers': 'Impacted Chargers',
                'open_count': 'Open Incidents',
                'closed_count': 'Closed Incidents',
                'within_tat': 'Closed Within TAT',
                'breached_tat': 'Breached TAT',
                'tat_efficiency': 'SLA Adherence %'
            })
            top10_disp.drop(columns=['fault_label'], errors='ignore').to_excel(writer, sheet_name='Top 10 Overall Faults', index=False)
            ws_t10 = writer.sheets['Top 10 Overall Faults']

            # Embedded Native Stacked BarChart on Top 10 Sheet
            chart_t10 = BarChart()
            chart_t10.type = "bar"
            chart_t10.grouping = "stacked"
            chart_t10.overlap = 100
            chart_t10.title = "Top 10 Repetitive Fault Patterns & SLA Compliance"
            chart_t10.width = 18
            chart_t10.height = 12
            chart_t10.x_axis.title = "Incident Volume"
            chart_t10.y_axis.title = "Fault Sub-Type"
            # Cols in top10_disp: 1:Category, 2:Fault Sub-Type, 3:Total Incidents, 4:Impacted Sites, 5:Impacted Chargers, 6:Open Incidents, 7:Closed Incidents, 8:Closed Within TAT, 9:Breached TAT, 10:SLA Adherence %
            data_t10 = Reference(ws_t10, min_col=8, max_col=9, min_row=1, max_row=len(top10_disp) + 1)
            cats_t10 = Reference(ws_t10, min_col=2, min_row=2, max_row=len(top10_disp) + 1)
            chart_t10.add_data(data_t10, titles_from_data=True)
            chart_t10.set_categories(cats_t10)
            ws_t10.add_chart(chart_t10, "L2")

        if not top20_st_rep.empty:
            top20_disp = top20_st_rep.rename(columns={
                'stationName': 'Station Name',
                'zone': 'Zone',
                'zme': 'Assigned ZME',
                'issueType': 'Category',
                'issueSubType': 'Fault Sub-Type',
                'fault_count': 'Fault Count at Site',
                'open_count': 'Open Incidents',
                'closed_count': 'Closed Incidents',
                'within_tat': 'Closed Within TAT',
                'breached_tat': 'Breached TAT',
                'tat_efficiency': 'Station SLA %'
            })
            top20_disp.drop(columns=['station_fault_label'], errors='ignore').to_excel(writer, sheet_name='Top 20 Station Faults', index=False)
            ws_t20 = writer.sheets['Top 20 Station Faults']

            # Embedded Native Stacked BarChart on Top 20 Sheet
            chart_t20 = BarChart()
            chart_t20.type = "bar"
            chart_t20.grouping = "stacked"
            chart_t20.overlap = 100
            chart_t20.title = "Top Station Incident Hotspots"
            chart_t20.width = 18
            chart_t20.height = 14
            chart_t20.x_axis.title = "Fault Count at Site"
            chart_t20.y_axis.title = "Station Name"
            # Cols in top20_disp: 1:Station Name, 2:Zone, 3:ZME, 4:Cat, 5:Sub, 6:Fault Count, 7:Open, 8:Closed, 9:Within TAT, 10:Breached TAT, 11:Station SLA %
            data_t20 = Reference(ws_t20, min_col=9, max_col=10, min_row=1, max_row=len(top20_disp) + 1)
            cats_t20 = Reference(ws_t20, min_col=1, min_row=2, max_row=len(top20_disp) + 1)
            chart_t20.add_data(data_t20, titles_from_data=True)
            chart_t20.set_categories(cats_t20)
            ws_t20.add_chart(chart_t20, "M2")

        # Sheet 3: ZME Scorecard
        zme_table = compute_zme_issue_table(filtered_df)
        if not zme_table.empty:
            zme_disp = zme_table.rename(columns={
                'zme': 'ZME Engineer',
                'zone': 'Zone',
                'total': 'Registered Faults',
                'open': 'Open Faults',
                'closed': 'Closed Faults',
                'within': 'Closed Within TAT',
                'outside': 'Closed Without TAT',
                'overdue': 'Overdue Open',
                'critical': 'Critical Faults',
                'cm_efficiency': 'CM Efficiency %',
                'tat_efficiency': 'TAT SLA %'
            })
            zme_disp.to_excel(writer, sheet_name='ZME Scorecard', index=False)
            ws_zme = writer.sheets['ZME Scorecard']

            # Embedded Native BarChart on ZME Scorecard Sheet
            chart_zme = BarChart()
            chart_zme.type = "bar"
            chart_zme.grouping = "stacked"
            chart_zme.overlap = 100
            chart_zme.title = "ZME Field Engineer SLA Performance"
            chart_zme.width = 18
            chart_zme.height = 12
            chart_zme.x_axis.title = "Incident Volume"
            chart_zme.y_axis.title = "ZME Engineer"
            # Cols in zme_disp: 1:ZME, 2:Zone, 3:Registered, 4:Open, 5:Closed, 6:Within TAT, 7:Outside TAT, 8:Overdue, 9:Critical, 10:CM%, 11:SLA%
            data_zme = Reference(ws_zme, min_col=6, max_col=7, min_row=1, max_row=len(zme_disp) + 1)
            cats_zme = Reference(ws_zme, min_col=1, min_row=2, max_row=len(zme_disp) + 1)
            chart_zme.add_data(data_zme, titles_from_data=True)
            chart_zme.set_categories(cats_zme)
            ws_zme.add_chart(chart_zme, "M2")

        # Sheet 4: Zone Performance
        zone_table = compute_zone_issue_table(filtered_df)
        if not zone_table.empty:
            zone_disp = zone_table.rename(columns={
                'zone': 'Zone / Region',
                'total': 'Total Faults',
                'open': 'Open Faults',
                'closed': 'Closed Faults',
                'within': 'Closed Within TAT',
                'outside': 'Closed Without TAT',
                'overdue': 'Overdue Open',
                'zme_count': 'ZME Count',
                'station_count': 'Impacted Sites',
                'cm_efficiency': 'CM Efficiency %',
                'tat_efficiency': 'TAT SLA %'
            })
            zone_disp.to_excel(writer, sheet_name='Zone Performance', index=False)
            ws_zone = writer.sheets['Zone Performance']

            # Embedded Native Clustered BarChart on Zone Sheet
            chart_zone = BarChart()
            chart_zone.type = "col"
            chart_zone.grouping = "clustered"
            chart_zone.title = "Regional Zone Incident & SLA Performance"
            chart_zone.width = 16
            chart_zone.height = 10
            chart_zone.y_axis.title = "Ticket Count"
            chart_zone.x_axis.title = "Zone / Region"
            # Cols in zone_disp: 1:Zone, 2:Total, 3:Open, 4:Closed, 5:Within TAT, 6:Outside TAT, 7:Overdue, 8:ZME Count, 9:Impacted Sites, 10:CM%, 11:SLA%
            data_zone = Reference(ws_zone, min_col=2, max_col=5, min_row=1, max_row=len(zone_disp) + 1)
            cats_zone = Reference(ws_zone, min_col=1, min_row=2, max_row=len(zone_disp) + 1)
            chart_zone.add_data(data_zone, titles_from_data=True)
            chart_zone.set_categories(cats_zone)
            ws_zone.add_chart(chart_zone, "M2")

        # Sheet 5: Filtered Issues Dataset
        if not filtered_df.empty:
            nice_cols = [
                'issueId', 'ocppId', 'stationId', 'stationName', 'zone', 'state', 'zme',
                'status', 'severity', 'issueType', 'issueSubType', 'issueDate', 'resolutionDate',
                'tatDays', 'tatCompliance', 'ageOfIssue', 'chargerMake', 'segment', 'description', 'correctiveAction'
            ]
            export_cols = [c for c in nice_cols if c in filtered_df.columns]
            filtered_df[export_cols].to_excel(writer, sheet_name='Filtered Issue Records', index=False)

        # Sheet 6: SLA Breached & Overdue
        breached_df = filtered_df[filtered_df['_Is_Closed_Without_'] | filtered_df['_Is_Overdue_']]
        if not breached_df.empty:
            nice_cols = [
                'issueId', 'ocppId', 'stationId', 'stationName', 'zone', 'state', 'zme',
                'status', 'severity', 'issueType', 'issueSubType', 'issueDate', 'resolutionDate',
                'tatDays', 'tatCompliance', 'ageOfIssue', 'chargerMake', 'segment', 'description', 'correctiveAction'
            ]
            export_cols_b = [c for c in nice_cols if c in breached_df.columns]
            breached_df[export_cols_b].to_excel(writer, sheet_name='SLA Breached & Overdue', index=False)

    output.seek(0)
    return output.getvalue()


# ─── Modern Plotly Chart Helpers ──────────────────────────────────────────────

def _modern_layout(title="", subtitle=None, height=360):
    full_title = f"<b>{title}</b>"
    if subtitle:
        full_title += f"<br><span style='font-size:11px;color:#64748B;font-weight:500;'>{subtitle}</span>"

    return dict(
        title=dict(
            text=full_title,
            font=dict(size=14, color="#0F172A", family="Plus Jakarta Sans, Inter, sans-serif"),
            x=0.01,
            y=0.96
        ),
        margin=dict(l=24, r=24, t=60, b=24),
        height=height,
        font=dict(family="Inter, sans-serif", color="#334155"),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        hoverlabel=dict(
            bgcolor="#FFFFFF",
            font_size=12,
            font_family="Inter, sans-serif",
            bordercolor="#DC2626"
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.28,
            xanchor="center",
            x=0.5,
            font=dict(size=11, family="Inter, sans-serif", color="#475569")
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="#F1F5F9",
            zeroline=False,
            tickfont=dict(size=11, color="#64748B")
        ),
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            tickfont=dict(size=11, color="#64748B")
        )
    )


def plot_top10_overall_faults(df_top10):
    """Plot horizontal stacked bar of top 10 overall repetitive faults."""
    df_sorted = df_top10.sort_values('total_count', ascending=True).copy()

    fig = go.Figure()

    fig.add_trace(go.Bar(
        y=df_sorted['fault_label'],
        x=df_sorted['within_tat'],
        name='Closed Within TAT',
        orientation='h',
        marker=dict(color='#10B981', line=dict(color='#059669', width=1)),
        hovertemplate='<b>%{y}</b><br>🟢 Closed Within TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_sorted['fault_label'],
        x=df_sorted['breached_tat'],
        name='Closed Breached TAT',
        orientation='h',
        marker=dict(color='#EF4444', line=dict(color='#DC2626', width=1)),
        hovertemplate='<b>%{y}</b><br>🔴 Closed Breached TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_sorted['fault_label'],
        x=df_sorted['open_count'],
        name='Active Open',
        orientation='h',
        marker=dict(color='#F59E0B', line=dict(color='#D97706', width=1)),
        hovertemplate='<b>%{y}</b><br>🟡 Active Open: <b>%{x:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="🏆 Top 10 Overall Repetitive Faults (Network-Wide)",
        subtitle="Ranked by incident frequency across all stations and equipment categories",
        height=max(400, len(df_sorted) * 40)
    )
    layout.update(dict(
        barmode='stack',
        xaxis=dict(title="Incident Volume", showgrid=True, gridcolor="#F1F5F9"),
        yaxis=dict(title=None, showgrid=False, autorange=True)
    ))
    fig.update_layout(**layout)
    return fig


def plot_top20_station_faults(df_top20):
    """Plot horizontal stacked bar of top 20 repetitive faults by station."""
    df_sorted = df_top20.sort_values('fault_count', ascending=True).copy()

    fig = go.Figure()

    fig.add_trace(go.Bar(
        y=df_sorted['station_fault_label'],
        x=df_sorted['within_tat'],
        name='Closed Within TAT',
        orientation='h',
        marker=dict(color='#10B981', line=dict(color='#059669', width=1)),
        hovertemplate='<b>%{y}</b><br>🟢 Within TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_sorted['station_fault_label'],
        x=df_sorted['breached_tat'],
        name='Breached TAT',
        orientation='h',
        marker=dict(color='#EF4444', line=dict(color='#DC2626', width=1)),
        hovertemplate='<b>%{y}</b><br>🔴 Breached TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_sorted['station_fault_label'],
        x=df_sorted['open_count'],
        name='Active Open',
        orientation='h',
        marker=dict(color='#F59E0B', line=dict(color='#D97706', width=1)),
        hovertemplate='<b>%{y}</b><br>🟡 Active Open: <b>%{x:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="📍 Top 20 Repetitive Faults by Stations (Station Hotspots)",
        subtitle="Specific site and failure combinations experiencing high recurring incidents",
        height=max(540, len(df_sorted) * 27)
    )
    layout.update(dict(
        barmode='stack',
        xaxis=dict(title="Incident Count at Site", showgrid=True, gridcolor="#F1F5F9"),
        yaxis=dict(title=None, showgrid=False, autorange=True)
    ))
    fig.update_layout(**layout)
    return fig


def plot_zme_sla_leaderboard(zme_df, top_n=10):
    """Plot horizontal interactive bar chart for top ZMEs with SLA compliance %."""
    df_top = zme_df.head(top_n).copy().sort_values('total', ascending=True)

    fig = go.Figure()

    fig.add_trace(go.Bar(
        y=df_top['zme'],
        x=df_top['within'],
        name='Closed Within TAT',
        orientation='h',
        marker=dict(color='#10B981', line=dict(color='#059669', width=1)),
        hovertemplate='<b>%{y}</b><br>🟢 Closed Within TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_top['zme'],
        x=df_top['outside'],
        name='Closed Breached TAT',
        orientation='h',
        marker=dict(color='#EF4444', line=dict(color='#DC2626', width=1)),
        hovertemplate='<b>%{y}</b><br>🔴 Closed Breached TAT: <b>%{x:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Bar(
        y=df_top['zme'],
        x=df_top['open'],
        name='Open / Pending',
        orientation='h',
        marker=dict(color='#F59E0B', line=dict(color='#D97706', width=1)),
        hovertemplate='<b>%{y}</b><br>🟡 Open Tickets: <b>%{x:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title=f"Top {min(top_n, len(df_top))} ZME Performance & SLA Compliance Breakdown",
        subtitle="Resolution status breakdown and SLA adherence by Field Engineer",
        height=max(380, len(df_top) * 38)
    )
    layout.update(dict(
        barmode='stack',
        xaxis=dict(title=None, showgrid=True, gridcolor="#F1F5F9"),
        yaxis=dict(title=None, showgrid=False, autorange=True)
    ))
    fig.update_layout(**layout)
    return fig


def plot_zone_efficiency_comparison(zone_df):
    """Plot dual-axis comparative bar for Zone Performance & Resolution Rates."""
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    fig.add_trace(
        go.Bar(
            x=zone_df['zone'],
            y=zone_df['total'],
            name='Total Registered',
            marker=dict(color='#991B1B', line=dict(color='#7F1D1D', width=1)),
            text=zone_df['total'],
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>Registered: <b>%{y:,}</b><extra></extra>'
        ),
        secondary_y=False
    )

    fig.add_trace(
        go.Bar(
            x=zone_df['zone'],
            y=zone_df['within'],
            name='Closed Within TAT',
            marker=dict(color='#10B981', line=dict(color='#059669', width=1)),
            text=zone_df['within'],
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>Within TAT: <b>%{y:,}</b><extra></extra>'
        ),
        secondary_y=False
    )

    fig.add_trace(
        go.Scatter(
            x=zone_df['zone'],
            y=zone_df['tat_efficiency'],
            name='SLA Compliance %',
            mode='lines+markers+text',
            line=dict(color='#2563EB', width=3, shape='spline'),
            marker=dict(size=9, color='#2563EB', symbol='circle', line=dict(color='#FFFFFF', width=2)),
            text=[f"{v:.2f}%" for v in zone_df['tat_efficiency']],
            textposition='top center',
            textfont=dict(size=11, color='#1E40AF', family="Inter, sans-serif"),
            hovertemplate='<b>%{x}</b><br>⚡ SLA Compliance: <b>%{y:.2f}%</b><extra></extra>'
        ),
        secondary_y=True
    )

    layout = _modern_layout(
        title="Zone Performance & SLA Compliance %",
        subtitle="Regional fault volume vs SLA achievement percentage",
        height=380
    )
    layout.update(dict(
        barmode='group',
        legend=dict(orientation="h", y=-0.25, x=0.5, xanchor="center")
    ))
    fig.update_layout(**layout)
    fig.update_yaxes(title_text="Ticket Count", secondary_y=False, showgrid=True, gridcolor="#F1F5F9")
    fig.update_yaxes(title_text="SLA %", secondary_y=True, range=[0, 115], showgrid=False)
    return fig


def plot_fault_trend_timeline(df):
    """Plot monthly fault volume trend and closure rate with smooth splines."""
    if df.empty or 'issueDate' not in df.columns:
        return None

    dt_df = df[df['issueDate'].notna()].copy()
    if dt_df.empty:
        return None

    dt_df['Period'] = dt_df['issueDate'].dt.strftime('%b %Y')
    dt_df['Period_Sort'] = dt_df['issueDate'].dt.to_period('M')

    trend_agg = dt_df.groupby(['Period_Sort', 'Period']).agg(
        Registered=('issueId', 'count'),
        Within_TAT=('_Is_Closed_Within_', 'sum'),
        Closed=('_Is_Closed_', 'sum'),
        Breached=('_Is_Closed_Without_', 'sum')
    ).reset_index().sort_values('Period_Sort')

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=trend_agg['Period'],
        y=trend_agg['Registered'],
        name='Faults Registered',
        marker=dict(color='#CBD5E1', line=dict(color='#94A3B8', width=1)),
        hovertemplate='<b>%{x}</b><br>Registered: <b>%{y:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=trend_agg['Period'],
        y=trend_agg['Within_TAT'],
        name='Resolved Within TAT',
        mode='lines+markers',
        line=dict(color='#10B981', width=3.5, shape='spline'),
        marker=dict(size=8, color='#10B981', line=dict(color='#FFFFFF', width=1.5)),
        hovertemplate='<b>%{x}</b><br>🟢 Resolved Within TAT: <b>%{y:,}</b><extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=trend_agg['Period'],
        y=trend_agg['Breached'],
        name='Resolved Breached TAT',
        mode='lines+markers',
        line=dict(color='#EF4444', width=2.5, dash='dot', shape='spline'),
        marker=dict(size=7, color='#EF4444', line=dict(color='#FFFFFF', width=1.5)),
        hovertemplate='<b>%{x}</b><br>🔴 Breached: <b>%{y:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="Monthly Operational Fault Trend & Resolution Adherence",
        subtitle="Timeline progression of incoming faults vs SLA-compliant closures",
        height=360
    )
    layout.update(dict(
        barmode='overlay',
        xaxis=dict(title=None, showgrid=False),
        yaxis=dict(title="Count", showgrid=True, gridcolor="#F1F5F9")
    ))
    fig.update_layout(**layout)
    return fig


def plot_donut_chart(labels, values, title, subtitle=None, colors=None, center_text=None):
    """Render modern donut chart with custom center text."""
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.64,
        marker=dict(colors=colors if colors else PALETTE_VIBRANT, line=dict(color='#FFFFFF', width=2.5)),
        textinfo='percent+value',
        textposition='outside',
        hovertemplate='<b>%{label}</b><br>Count: <b>%{value:,}</b> (%{percent})<extra></extra>',
        insidetextfont=dict(color='#FFFFFF', size=11, family='Inter, sans-serif')
    )])

    layout = _modern_layout(title=title, subtitle=subtitle, height=340)
    layout.update(dict(
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.30, xanchor="center", x=0.5)
    ))
    layout.pop('yaxis', None)
    layout.pop('xaxis', None)

    if center_text:
        layout['annotations'] = [dict(
            text=center_text,
            x=0.5, y=0.5,
            font_size=16,
            font_family='Plus Jakarta Sans, Inter, sans-serif',
            font_weight='bold',
            font_color='#0F172A',
            showarrow=False
        )]

    fig.update_layout(**layout)
    return fig


def plot_pareto_root_causes(df, top_n=8):
    """Pareto chart of top issue categories and cumulative percentage."""
    if df.empty or 'issueType' not in df.columns:
        return None

    cat_counts = df['issueType'].value_counts().reset_index()
    cat_counts.columns = ['Category', 'Count']
    cat_counts = cat_counts.head(top_n)

    cat_counts['Cumulative'] = cat_counts['Count'].cumsum()
    tot = cat_counts['Count'].sum()
    cat_counts['Cum_Pct'] = (cat_counts['Cumulative'] / tot * 100).round(2)

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    fig.add_trace(
        go.Bar(
            x=cat_counts['Category'],
            y=cat_counts['Count'],
            name='Incident Count',
            marker=dict(color='#DC2626', line=dict(color='#B91C1C', width=1)),
            text=cat_counts['Count'],
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>Incidents: <b>%{y:,}</b><extra></extra>'
        ),
        secondary_y=False
    )

    fig.add_trace(
        go.Scatter(
            x=cat_counts['Category'],
            y=cat_counts['Cum_Pct'],
            name='Cumulative Share %',
            mode='lines+markers',
            line=dict(color='#2563EB', width=2.8, shape='spline'),
            marker=dict(size=8, color='#2563EB', line=dict(color='#FFFFFF', width=1.5)),
            hovertemplate='<b>%{x}</b><br>Cumulative: <b>%{y:.2f}%</b><extra></extra>'
        ),
        secondary_y=True
    )

    layout = _modern_layout(
        title=f"Top {min(top_n, len(cat_counts))} Root Cause Categories (Pareto Analysis)",
        subtitle="Most prevalent equipment and infrastructure fault categories",
        height=360
    )
    layout.update(dict(
        legend=dict(orientation="h", y=-0.26, x=0.5, xanchor="center")
    ))
    fig.update_layout(**layout)
    fig.update_yaxes(title_text="Faults", secondary_y=False, showgrid=True, gridcolor="#F1F5F9")
    fig.update_yaxes(title_text="Cumulative %", secondary_y=True, range=[0, 105], showgrid=False)
    return fig


def plot_charger_make_analysis(df):
    """Plot fault count distribution across charger OEMs / Makes for Charger & Charger_Software issues only."""
    if df.empty or 'chargerMake' not in df.columns:
        return None

    make_df = df[df['chargerMake'].notna() & (df['chargerMake'] != 'Unknown') & (df['chargerMake'] != '') & (df['chargerMake'] != 'None')].copy()
    
    # Filter ONLY for Charger & Charger_Software issue types
    if 'issueType' in make_df.columns:
        valid_types = ['Charger', 'Charger Software', 'Charger_Software']
        make_df = make_df[make_df['issueType'].astype(str).str.strip().isin(valid_types)]

    if make_df.empty:
        return None

    # Normalize charger OEM names (e.g. EXICOM -> Exicom)
    make_df['chargerMake'] = make_df['chargerMake'].astype(str).str.strip().str.title()

    make_agg = make_df.groupby('chargerMake').agg(
        Total=('issueId', 'count'),
        Within_TAT=('_Is_Closed_Within_', 'sum'),
        Breached=('_Is_Closed_Without_', 'sum'),
        Open=('_Is_Open_', 'sum')
    ).reset_index().sort_values('Total', ascending=False).head(8)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=make_agg['chargerMake'],
        y=make_agg['Within_TAT'],
        name='Within TAT',
        marker=dict(color='#10B981', line=dict(color='#059669', width=1)),
        hovertemplate='<b>%{x}</b><br>🟢 Within TAT: <b>%{y:,}</b><extra></extra>'
    ))
    fig.add_trace(go.Bar(
        x=make_agg['chargerMake'],
        y=make_agg['Breached'],
        name='Breached TAT',
        marker=dict(color='#EF4444', line=dict(color='#DC2626', width=1)),
        hovertemplate='<b>%{x}</b><br>🔴 Breached TAT: <b>%{y:,}</b><extra></extra>'
    ))
    fig.add_trace(go.Bar(
        x=make_agg['chargerMake'],
        y=make_agg['Open'],
        name='Open Faults',
        marker=dict(color='#F59E0B', line=dict(color='#D97706', width=1)),
        hovertemplate='<b>%{x}</b><br>🟡 Open: <b>%{y:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="Charger OEM Reliability & SLA Adherence",
        subtitle="Fault volume and resolution breakdown for Charger & Charger_Software issues by OEM",
        height=360
    )
    layout.update(dict(
        barmode='stack',
        xaxis=dict(title=None, showgrid=False),
        yaxis=dict(title="Fault Count", showgrid=True, gridcolor="#F1F5F9")
    ))
    fig.update_layout(**layout)
    return fig


def plot_station_hotspots(df, top_n=10):
    """Plot horizontal bar of top incident stations."""
    if df.empty or 'stationName' not in df.columns:
        return None

    st_counts = df[df['stationName'] != 'N/A']['stationName'].value_counts().reset_index()
    st_counts.columns = ['Station', 'Faults']
    st_top = st_counts.head(top_n).sort_values('Faults', ascending=True)

    fig = go.Figure(go.Bar(
        x=st_top['Faults'],
        y=st_top['Station'],
        orientation='h',
        marker=dict(color='#991B1B', line=dict(color='#7F1D1D', width=1)),
        text=st_top['Faults'],
        textposition='outside',
        textfont=dict(size=11, family='Inter, sans-serif', color='#0F172A'),
        hovertemplate='<b>%{y}</b><br>Faults Logged: <b>%{x:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title=f"Top {min(top_n, len(st_top))} Station Incident Hotspots",
        subtitle="Sites with the highest volume of registered operational issues",
        height=max(360, len(st_top) * 35)
    )
    layout.update(dict(
        xaxis=dict(title=None, showgrid=True, gridcolor="#F1F5F9"),
        yaxis=dict(title=None, showgrid=False, autorange=True)
    ))
    fig.update_layout(**layout)
    return fig


def plot_top_issue_types_barchart(df_types):
    """Plot vertical bar chart for Top 10 Overall Repetitive Issue Types."""
    if df_types.empty:
        return None

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_types['Issue Type'],
        y=df_types['Total_Faults'],
        name='Total Faults',
        marker=dict(
            color='#991B1B',
            line=dict(color='#7F1D1D', width=1)
        ),
        text=df_types['Total_Faults'],
        textposition='outside',
        textfont=dict(size=11, weight='bold', color='#1E293B'),
        hovertemplate='<b>%{x}</b><br>Total Faults: <b>%{y:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="Top 10 Overall Repetitive Issue Types",
        subtitle="Network-wide fault count aggregated by equipment category",
        height=350
    )
    layout.update(dict(
        xaxis=dict(title="issueType", showgrid=False, tickangle=-15),
        yaxis=dict(title="Total_Faults", showgrid=True, gridcolor="#F1F5F9")
    ))
    fig.update_layout(**layout)
    return fig


def plot_top_subtypes_barchart(df_subtypes, cat_name):
    """Plot vertical bar chart for Top 5 Sub-Types of a selected Issue Type or Station."""
    if df_subtypes.empty:
        return None

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_subtypes['Issue Sub-Type'],
        y=df_subtypes['Subtype_Count'],
        name='Sub-Type Count',
        marker=dict(
            color='#DC2626',
            line=dict(color='#B91C1C', width=1)
        ),
        text=df_subtypes['Subtype_Count'],
        textposition='outside',
        textfont=dict(size=11, weight='bold', color='#1E293B'),
        hovertemplate=f'<b>{cat_name}</b> - %{{x}}<br>Occurrences: <b>%{{y:,}}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title=f"Top 5 Sub-Types for '{str(cat_name)[:28]}'",
        subtitle=f"Most frequent failure sub-types under {str(cat_name)[:28]}",
        height=320
    )
    layout.update(dict(
        xaxis=dict(title="IssueSubType", showgrid=False, tickangle=-15),
        yaxis=dict(title="Subtype_Count", showgrid=True, gridcolor="#F1F5F9")
    ))
    fig.update_layout(**layout)
    return fig


def plot_top_stations_barchart(df_top_stations):
    """Plot vertical bar chart for Top Station Hotspots."""
    if df_top_stations.empty:
        return None

    df_top10 = df_top_stations.head(10).copy()
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_top10['stationName'].apply(lambda s: str(s)[:20] + '...' if len(str(s)) > 22 else str(s)),
        y=df_top10['fault_count'],
        name='Fault Count',
        marker=dict(
            color='#DC2626',
            line=dict(color='#B91C1C', width=1)
        ),
        text=df_top10['fault_count'],
        textposition='outside',
        textfont=dict(size=11, weight='bold', color='#1E293B'),
        hovertemplate='<b>%{x}</b><br>Faults at Site: <b>%{y:,}</b><extra></extra>'
    ))

    layout = _modern_layout(
        title="Top 10 Station Hotspots Overview",
        subtitle="Site locations with highest total repetitive fault occurrences",
        height=350
    )
    layout.update(dict(
        xaxis=dict(title="Station Location", showgrid=False, tickangle=-15),
        yaxis=dict(title="Fault Count", showgrid=True, gridcolor="#F1F5F9")
    ))
    fig.update_layout(**layout)
    return fig


def plot_subtype_occurrence_linechart(df, group_col='issueSubType', title="Sub-type Occurrence Trend", subtitle=None):
    """Plot multi-line timeline chart showing occurrence trends for each fault sub-type over time."""
    if df.empty or 'issueDate' not in df.columns or group_col not in df.columns:
        return None

    dt_df = df[df['issueDate'].notna()].copy()
    if dt_df.empty:
        return None

    dt_df['Period'] = dt_df['issueDate'].dt.strftime('%b %Y')
    dt_df['Period_Sort'] = dt_df['issueDate'].dt.to_period('M')

    # Aggregate occurrences by Month and Sub-type
    trend_df = dt_df.groupby(['Period_Sort', 'Period', group_col]).size().reset_index(name='Occurrences')
    trend_df = trend_df.sort_values('Period_Sort')

    if trend_df.empty:
        return None

    fig = go.Figure()
    top_subtypes = dt_df[group_col].value_counts().head(8).index.tolist()
    filtered_trend = trend_df[trend_df[group_col].isin(top_subtypes)]

    palette = [
        '#DC2626', '#2563EB', '#10B981', '#F59E0B', '#8B5CF6',
        '#EC4899', '#06B6D4', '#64748B', '#D97706', '#059669'
    ]

    for idx, subtype in enumerate(top_subtypes):
        sub_slice = filtered_trend[filtered_trend[group_col] == subtype]
        if sub_slice.empty:
            continue
        color = palette[idx % len(palette)]
        fig.add_trace(go.Scatter(
            x=sub_slice['Period'],
            y=sub_slice['Occurrences'],
            name=str(subtype),
            mode='lines+markers+text',
            text=sub_slice['Occurrences'],
            textposition='top center',
            textfont=dict(size=10, family="Inter, sans-serif"),
            line=dict(color=color, width=3, shape='spline'),
            marker=dict(size=7, color=color, line=dict(color='#FFFFFF', width=1.5)),
            hovertemplate=f'<b>{subtype}</b><br>Period: <b>%{{x}}</b><br>Occurrences: <b>%{{y:,}}</b><extra></extra>'
        ))

    layout = _modern_layout(
        title=title,
        subtitle=subtitle if subtitle else "Timeline frequency & occurrence progression by sub-type",
        height=380
    )
    layout.update(dict(
        xaxis=dict(title=None, showgrid=False),
        yaxis=dict(title="Occurrence Count", showgrid=True, gridcolor="#F1F5F9", dtick=1),
        legend=dict(orientation="h", yanchor="bottom", y=-0.34, xanchor="center", x=0.5)
    ))
    fig.update_layout(**layout)
    return fig


# ─── Executive CSS & Styling ──────────────────────────────────────────────────

def inject_modern_css():
    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=Plus+Jakarta+Sans:wght@600;700;800&display=swap');

        /* Global Reset & Executive Theme */
        .stApp, [data-testid="stMain"], [data-testid="stHeader"] {
            background-color: #F8FAFC !important;
            font-family: 'Inter', sans-serif !important;
        }

        [data-testid="stSidebar"] {
            display: none !important;
        }

        html, body, p, span, label, div {
            font-family: 'Inter', sans-serif;
            color: #1E293B;
        }

        h1, h2, h3, h4, h5, h6 {
            font-family: 'Plus Jakarta Sans', 'Inter', sans-serif !important;
            color: #0F172A !important;
            font-weight: 800 !important;
            letter-spacing: -0.025em !important;
        }

        /* ── Full-Width Executive Header Banner ── */
        .exec-banner {
            background: linear-gradient(135deg, #3B0707 0%, #7F1D1D 35%, #991B1B 70%, #DC2626 100%);
            padding: 2.0rem 2.6rem;
            border-radius: 24px;
            color: #FFFFFF !important;
            box-shadow: 0 16px 40px rgba(153,27,27,0.22), inset 0 1px 0 rgba(255,255,255,0.2);
            margin-bottom: 1.4rem;
            border-left: 8px solid #F87171;
            position: relative;
            overflow: hidden;
        }
        .exec-banner::after {
            content: '';
            position: absolute;
            top: -40%;
            right: -8%;
            width: 420px;
            height: 420px;
            background: radial-gradient(circle, rgba(255,255,255,0.15) 0%, rgba(255,255,255,0) 70%);
            border-radius: 50%;
            pointer-events: none;
        }
        .exec-banner * {
            color: #FFFFFF !important;
        }
        .exec-badge {
            background: rgba(255,255,255,0.20);
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
            color: #FFFFFF !important;
            font-size: 0.74rem;
            font-weight: 800;
            padding: 6px 16px;
            border-radius: 20px;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            display: inline-block;
            margin-bottom: 0.55rem;
            border: 1px solid rgba(255,255,255,0.40);
        }
        .exec-title {
            font-size: 2.25rem !important;
            font-weight: 900 !important;
            color: #FFFFFF !important;
            margin: 0.15rem 0 0.40rem 0 !important;
            line-height: 1.2 !important;
            letter-spacing: -0.03em !important;
            text-shadow: 0 3px 12px rgba(0,0,0,0.30) !important;
        }
        .exec-subtitle {
            font-size: 0.98rem !important;
            color: #FEE2E2 !important;
            margin-top: 0.2rem !important;
            font-weight: 500 !important;
            opacity: 0.96;
        }

        /* ── Streamlit Pills Navigation ── */
        div[data-testid="stPills"],
        div[data-testid="stSegmentedControl"],
        div[data-testid="stButtonGroupRoot"] {
            display: flex !important;
            justify-content: flex-start !important;
            gap: 0.65rem !important;
            margin-bottom: 1.5rem !important;
            flex-wrap: wrap !important;
        }
        div[data-testid="stPills"] button,
        div[data-testid="stSegmentedControl"] button,
        div[data-testid="stButtonGroupRoot"] button {
            border-radius: 50px !important;
            padding: 0.65rem 1.55rem !important;
            font-size: 0.93rem !important;
            font-weight: 700 !important;
            border: 1.5px solid #E2E8F0 !important;
            background: #FFFFFF !important;
            color: #475569 !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.03) !important;
            transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        }
        div[data-testid="stPills"] button:hover,
        div[data-testid="stSegmentedControl"] button:hover,
        div[data-testid="stButtonGroupRoot"] button:hover {
            border-color: #FCA5A5 !important;
            color: #991B1B !important;
            background: #FFF5F5 !important;
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 18px rgba(153,27,27,0.10) !important;
        }
        div[data-testid="stPills"] button[aria-pressed="true"],
        div[data-testid="stPills"] button[data-selected="true"],
        div[data-testid="stSegmentedControl"] button[aria-pressed="true"],
        div[data-testid="stSegmentedControl"] button[data-selected="true"],
        div[data-testid="stButtonGroupRoot"] button[aria-pressed="true"],
        div[data-testid="stButtonGroupRoot"] button[data-selected="true"],
        button[data-testid="stBaseButton-pillsActive"] {
            background: linear-gradient(135deg, #991B1B 0%, #DC2626 100%) !important;
            color: #FFFFFF !important;
            border-color: #DC2626 !important;
            box-shadow: 0 6px 20px rgba(220,38,38,0.38) !important;
        }
        div[data-testid="stPills"] button[aria-pressed="true"] *,
        div[data-testid="stPills"] button[data-selected="true"] *,
        button[data-testid="stBaseButton-pillsActive"] * {
            color: #FFFFFF !important;
        }

        /* ── Modern Streamlit Tabs Styling ── */
        div[data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 8px !important;
            border-bottom: 2px solid #E2E8F0 !important;
            padding-bottom: 4px !important;
        }
        div[data-testid="stTabs"] [data-baseweb="tab"] {
            height: 44px !important;
            white-space: pre !important;
            border-radius: 12px 12px 0 0 !important;
            font-weight: 700 !important;
            font-size: 0.92rem !important;
            color: #64748B !important;
            padding: 0 16px !important;
            transition: all 0.2s ease !important;
        }
        div[data-testid="stTabs"] [aria-selected="true"] {
            color: #991B1B !important;
            border-bottom: 3px solid #DC2626 !important;
            background: #FFF5F5 !important;
        }

        /* ── Modern KPI Cards ── */
        .kpi-grid {
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 0.9rem;
            margin-bottom: 1.6rem;
        }
        @media (max-width: 1400px) {
            .kpi-grid { grid-template-columns: repeat(4, 1fr); }
        }
        @media (max-width: 900px) {
            .kpi-grid { grid-template-columns: repeat(2, 1fr); }
        }
        .metric-card {
            background: #FFFFFF;
            border: 1px solid #F1F5F9;
            border-radius: 18px;
            padding: 1.20rem 1.30rem;
            box-shadow: 0 6px 20px rgba(15,23,42,0.04);
            transition: all 0.28s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }
        .metric-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
        }
        .metric-card.red::before     { background: linear-gradient(90deg, #DC2626, #EF4444); }
        .metric-card.darkred::before { background: linear-gradient(90deg, #7F1D1D, #991B1B); }
        .metric-card.amber::before   { background: linear-gradient(90deg, #D97706, #F59E0B); }
        .metric-card.green::before   { background: linear-gradient(90deg, #059669, #10B981); }
        .metric-card.blue::before    { background: linear-gradient(90deg, #1D4ED8, #3B82F6); }
        .metric-card.purple::before  { background: linear-gradient(90deg, #6D28D9, #8B5CF6); }
        .metric-card.slate::before   { background: linear-gradient(90deg, #475569, #64748B); }

        .metric-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 14px 32px rgba(153,27,27,0.10);
            border-color: #FCA5A5;
        }
        .metric-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 0.4rem;
        }
        .metric-icon {
            font-size: 1.65rem;
            line-height: 1;
        }
        .metric-label {
            font-size: 0.72rem;
            font-weight: 800;
            color: #64748B !important;
            text-transform: uppercase;
            letter-spacing: 0.09em;
        }
        .metric-val {
            font-size: 1.90rem;
            font-weight: 900;
            color: #0F172A !important;
            line-height: 1.15;
            letter-spacing: -0.025em;
            margin: 0.2rem 0;
            font-family: 'Plus Jakarta Sans', sans-serif;
        }
        .metric-sub {
            font-size: 0.76rem;
            color: #64748B !important;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 4px;
        }
        .metric-badge {
            font-size: 0.70rem;
            font-weight: 700;
            padding: 3px 10px;
            border-radius: 12px;
            display: inline-block;
        }
        .metric-badge.green { background: #DCFCE7; color: #15803D !important; }
        .metric-badge.amber { background: #FEF3C7; color: #B45309 !important; }
        .metric-badge.red   { background: #FEE2E2; color: #B91C1C !important; }

        /* ── Modern Section Header ── */
        .section-header {
            font-size: 1.25rem;
            font-weight: 800;
            color: #0F172A !important;
            margin-top: 1.6rem;
            margin-bottom: 1.1rem;
            padding-bottom: 0.55rem;
            border-bottom: 2px solid #F1F5F9;
            display: flex;
            align-items: center;
            gap: 8px;
            font-family: 'Plus Jakarta Sans', sans-serif;
        }
        .section-header-accent {
            color: #DC2626 !important;
        }

        /* ── Buttons & Downloads ── */
        .stButton > button, div[data-testid="stDownloadButton"] > button {
            background: linear-gradient(135deg, #801B1B 0%, #B91C1C 50%, #DC2626 100%) !important;
            color: #FFFFFF !important;
            border-radius: 12px !important;
            font-weight: 700 !important;
            font-size: 0.90rem !important;
            border: none !important;
            padding: 0.65rem 1.45rem !important;
            box-shadow: 0 4px 16px rgba(185,28,28,0.25) !important;
            transition: all 0.22s ease !important;
        }
        .stButton > button:hover, div[data-testid="stDownloadButton"] > button:hover {
            background: linear-gradient(135deg, #6B1111 0%, #991B1B 50%, #B91C1C 100%) !important;
            box-shadow: 0 8px 24px rgba(153,27,27,0.38) !important;
            transform: translateY(-2px) !important;
        }

        /* ── Chart Container Cards ── */
        .chart-card {
            background: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 20px;
            padding: 1.2rem;
            box-shadow: 0 4px 20px rgba(15,23,42,0.03);
            margin-bottom: 1.2rem;
            transition: all 0.25s ease;
        }
        .chart-card:hover {
            box-shadow: 0 10px 28px rgba(15,23,42,0.07);
            border-color: #CBD5E1;
        }

        /* ── Container Cards ── */
        .content-card {
            background: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 20px;
            padding: 1.5rem 1.8rem;
            box-shadow: 0 4px 20px rgba(0,0,0,0.03);
            margin-bottom: 1.3rem;
        }

        /* ── Dataframe Styling ── */
        [data-testid="stDataFrame"] {
            border-radius: 16px;
            overflow: hidden;
            border: 1px solid #E2E8F0;
            box-shadow: 0 4px 16px rgba(0,0,0,0.03);
        }

        /* ── Upload Area ── */
        .upload-card {
            background: #FFFFFF;
            border: 2px dashed #FCA5A5;
            border-radius: 20px;
            padding: 2.4rem 2.0rem;
            text-align: center;
            transition: all 0.25s ease;
        }
        .upload-card:hover {
            border-color: #DC2626;
            transform: translateY(-2px);
            box-shadow: 0 10px 30px rgba(153,27,27,0.10);
        }
        </style>
    """, unsafe_allow_html=True)


# ─── Navigation Header ────────────────────────────────────────────────────────

NAV_ITEMS = [
    "📊 Executive Analytics",
    "🔁 Repetitive Faults",
    "🚨 Critical & SLA Breached",
    "👷 ZME & Zone Governance",
    "📋 Issue Explorer",
    "📤 Data Upload & Validation"
]

def render_header_and_nav():
    """Render the executive header banner and pills navigation bar."""
    if 'active_nav' not in st.session_state:
        st.session_state.active_nav = NAV_ITEMS[0]

    st.markdown("""
        <div class="exec-banner">
            <div>
                <span class="exec-badge">⚡ CHARGEZONE OPERATIONS • SLA GOVERNANCE</span>
                <div class="exec-title">MPR - Issue Tracker Dashboard</div>
                <div class="exec-subtitle">Real-time incident tracking, turnaround time (TAT) compliance, and field engineering analytics.</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    selected_nav = st.pills(
        "Section Navigation",
        NAV_ITEMS,
        default=st.session_state.active_nav,
        label_visibility="collapsed",
        key="top_navbar_pills"
    )

    if selected_nav and selected_nav != st.session_state.active_nav:
        st.session_state.active_nav = selected_nav
        st.rerun()

    return st.session_state.active_nav


# ─── Filter Toolbar ───────────────────────────────────────────────────────────

def render_filter_toolbar(issue_df):
    """Render clean horizontal filter toolbar with Month, Zone, ZME, Severity, Search, and Excel Download."""
    all_months = get_available_months(issue_df)

    with st.container():
        st.markdown("""
            <div style="background: #FFFFFF; border: 1.5px solid #FEE2E2; border-radius: 16px;
                        padding: 0.75rem 1.4rem; margin-bottom: 1.1rem; box-shadow: 0 2px 10px rgba(153,27,27,0.03);
                        display: flex; align-items: center; justify-content: space-between;">
                <span style="color: #991B1B; font-weight: 800; font-size: 0.84rem; text-transform: uppercase; letter-spacing: 0.08em; display: flex; align-items: center; gap: 6px;">
                    ⚡ Governance Filters &amp; Report Export
                </span>
                <span style="font-size: 0.80rem; color: #64748B; font-weight: 600;">
                    Live Incident Pipeline
                </span>
            </div>
        """, unsafe_allow_html=True)

        col_m, col_st, col_z, col_zme = st.columns([2.5, 2.5, 2.5, 2.5])

        with col_m:
            selected_months = st.multiselect(
                "🗓️ Months:",
                all_months,
                default=all_months,
                key="flt_months"
            )
            if not selected_months:
                selected_months = all_months

        # Base filtered slice for dynamic State/Zone/ZME dropdowns
        temp_df = issue_df.copy()
        if selected_months:
            temp_df = temp_df[temp_df['month'].isin(selected_months)]

        # Available States
        available_states = sorted(list(temp_df['state'].dropna().unique())) if not temp_df.empty else []
        with col_st:
            selected_states = st.multiselect("🏛️ State / Location:", available_states, default=available_states, key="flt_states")
            if not selected_states:
                selected_states = available_states

        # Available Zones within selected states
        if selected_states and not temp_df.empty:
            zone_pool_df = temp_df[temp_df['state'].isin(selected_states)]
        else:
            zone_pool_df = temp_df

        available_zones = sorted(list(zone_pool_df['zone'].dropna().unique())) if not zone_pool_df.empty else []
        with col_z:
            selected_zones = st.multiselect("🏢 Zone / Region:", available_zones, default=available_zones, key="flt_zones")
            if not selected_zones:
                selected_zones = available_zones

        # ZMEs within selected zones
        if selected_zones and not zone_pool_df.empty:
            zme_pool = sorted(list(zone_pool_df[zone_pool_df['zone'].isin(selected_zones)]['zme'].dropna().unique()))
        else:
            zme_pool = sorted(list(temp_df['zme'].dropna().unique())) if not temp_df.empty else []

        with col_zme:
            selected_zmes = st.multiselect("👷 ZME / Engineer:", zme_pool, default=zme_pool, key="flt_zmes")
            if not selected_zmes:
                selected_zmes = zme_pool

        # Secondary filter row: Charger Make, Search, and Excel Export
        col_make, col_search, col_exp = st.columns([2.5, 5.0, 2.5])
        with col_make:
            available_makes = ['All'] + sorted(list(issue_df[issue_df['chargerMake'] != 'Unknown']['chargerMake'].dropna().unique())) if not issue_df.empty else ['All']
            selected_make = st.selectbox("🔌 Charger Make:", available_makes, index=0, key="flt_make")

        with col_search:
            search_query = st.text_input("🔍 Search (Issue ID, Station, Charger ID, Keyword):", placeholder="e.g. CZ-0010, KA0010, Ador, Canopy...", key="flt_search")

        # Apply ALL filters for exact export & display according to active user choices
        filtered_for_export = apply_filters(
            issue_df,
            months=selected_months,
            states=selected_states,
            zones=selected_zones,
            zmes=selected_zmes,
            make=selected_make,
            query=search_query
        )

        active_filters_meta = {
            'states': ", ".join(selected_states[:3]) if len(selected_states) <= 3 else f"{len(selected_states)} Selected",
            'zones': ", ".join(selected_zones[:3]) if len(selected_zones) <= 3 else f"{len(selected_zones)} Selected",
            'make': selected_make,
            'search': search_query if search_query else "None"
        }

        # Excel Report Download
        with col_exp:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            report_bytes = generate_issue_excel_report(filtered_for_export, issue_df, tuple(selected_months), active_filters_meta)
            month_slug = "_".join(selected_months[:3]) if len(selected_months) <= 3 else f"{selected_months[0]}_to_{selected_months[-1]}"
            st.download_button(
                label="📥 Export Report with Charts (.xlsx)",
                data=report_bytes,
                file_name=f"ChargeZone_Issue_Tracker_{month_slug}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="flt_download_report_btn"
            )

    return selected_months, selected_states, selected_zones, selected_zmes, selected_make, search_query


def apply_filters(df, months, states, zones, zmes, make, query):
    """Apply all active filters to the issue dataframe."""
    if df.empty:
        return df

    filtered = df.copy()
    if months:
        filtered = filtered[filtered['month'].isin(months)]
    if states:
        filtered = filtered[filtered['state'].isin(states)]
    if zones:
        filtered = filtered[filtered['zone'].isin(zones)]
    if zmes:
        filtered = filtered[filtered['zme'].isin(zmes)]
    if make and make != 'All':
        filtered = filtered[filtered['chargerMake'] == make]

    if query and str(query).strip():
        q = str(query).strip().lower()
        search_mask = (
            filtered['issueId'].astype(str).str.lower().str.contains(q) |
            filtered['stationId'].astype(str).str.lower().str.contains(q) |
            filtered['stationName'].astype(str).str.lower().str.contains(q) |
            filtered['ocppId'].astype(str).str.lower().str.contains(q) |
            filtered['description'].astype(str).str.lower().str.contains(q) |
            filtered['issueType'].astype(str).str.lower().str.contains(q) |
            filtered['issueSubType'].astype(str).str.lower().str.contains(q) |
            filtered['zme'].astype(str).str.lower().str.contains(q) |
            filtered['state'].astype(str).str.lower().str.contains(q)
        )
        filtered = filtered[search_mask]

    return filtered


# ─── View 1: Executive Analytics ──────────────────────────────────────────────

def render_executive_analytics(filtered_df, full_df):
    """Render primary KPI cards and high-level analytical Plotly charts."""
    st.markdown('<div class="section-header">📊 <span class="section-header-accent">Executive KPI</span> Summary</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.warning("⚠️ No issue records match the selected filter criteria. Try broadening your filter selections.")
        return

    # ── 7 KPI Metrics ────────────────────────────────────────────────────────
    total_issues = len(filtered_df)
    total_open = int(filtered_df['_Is_Open_'].sum())
    total_closed = int(filtered_df['_Is_Closed_'].sum())
    closed_within = int(filtered_df['_Is_Closed_Within_'].sum())
    closed_without = int(filtered_df['_Is_Closed_Without_'].sum())
    overdue_count = int(filtered_df['_Is_Overdue_'].sum())

    overall_cm_eff = (total_closed / total_issues * 100) if total_issues > 0 else 0.0
    cm_tat_eff = (closed_within / total_issues * 100) if total_issues > 0 else 0.0

    kpi_html = f"""
    <div class="kpi-grid">
        <div class="metric-card darkred">
            <div class="metric-header">
                <span class="metric-label">Registered</span>
                <span class="metric-icon">🗂️</span>
            </div>
            <div class="metric-val">{total_issues:,}</div>
            <div class="metric-sub">Total Logged Faults</div>
        </div>
        <div class="metric-card amber">
            <div class="metric-header">
                <span class="metric-label">Active Open</span>
                <span class="metric-icon">⏳</span>
            </div>
            <div class="metric-val">{total_open:,}</div>
            <div class="metric-sub">
                <span class="metric-badge {'red' if overdue_count > 0 else 'green'}">{overdue_count} Overdue</span>
            </div>
        </div>
        <div class="metric-card blue">
            <div class="metric-header">
                <span class="metric-label">Resolved / Closed</span>
                <span class="metric-icon">✅</span>
            </div>
            <div class="metric-val">{total_closed:,}</div>
            <div class="metric-sub">Total Addressed</div>
        </div>
        <div class="metric-card green">
            <div class="metric-header">
                <span class="metric-label">Within TAT</span>
                <span class="metric-icon">🟢</span>
            </div>
            <div class="metric-val">{closed_within:,}</div>
            <div class="metric-sub">SLA Compliant</div>
        </div>
        <div class="metric-card red">
            <div class="metric-header">
                <span class="metric-label">Breached TAT</span>
                <span class="metric-icon">🔴</span>
            </div>
            <div class="metric-val">{closed_without:,}</div>
            <div class="metric-sub">SLA Violations</div>
        </div>
        <div class="metric-card {'green' if overall_cm_eff >= 85 else 'amber'}">
            <div class="metric-header">
                <span class="metric-label">Overall CM %</span>
                <span class="metric-icon">📈</span>
            </div>
            <div class="metric-val">{overall_cm_eff:.2f}%</div>
            <div class="metric-sub">Closure Rate</div>
        </div>
        <div class="metric-card {'green' if cm_tat_eff >= 80 else 'amber'}">
            <div class="metric-header">
                <span class="metric-label">TAT SLA %</span>
                <span class="metric-icon">⚡</span>
            </div>
            <div class="metric-val">{cm_tat_eff:.2f}%</div>
            <div class="metric-sub">Within TAT / Logged</div>
        </div>
    </div>
    """
    st.markdown(kpi_html, unsafe_allow_html=True)

    # ── Charts Section 1: ZME Leaderboard & Zone Performance ─────────────────
    st.markdown('<div class="section-header">👷 <span class="section-header-accent">Field Engineering</span> & Regional Performance</div>', unsafe_allow_html=True)
    col_c1, col_c2 = st.columns([6, 6])

    with col_c1:
        zme_agg = compute_zme_issue_table(filtered_df)
        if not zme_agg.empty:
            st.plotly_chart(plot_zme_sla_leaderboard(zme_agg, top_n=10), use_container_width=True)
        else:
            st.info("No ZME breakdown data available.")

    with col_c2:
        zone_agg = compute_zone_issue_table(filtered_df)
        if not zone_agg.empty:
            st.plotly_chart(plot_zone_efficiency_comparison(zone_agg), use_container_width=True)
        else:
            st.info("No Zone comparison data available.")

    # ── Charts Section 2: Pipeline Donuts & Fault Trend ───────────────────────
    st.markdown('<div class="section-header">📌 <span class="section-header-accent">Status Pipeline</span> & Trend Analytics</div>', unsafe_allow_html=True)
    col_d1, col_d2, col_d3 = st.columns([4, 4, 4])

    with col_d1:
        status_counts = filtered_df['status'].value_counts()
        colors_status = [PALETTE_STATUS.get(s, '#64748B') for s in status_counts.index]
        fig_stat = plot_donut_chart(
            labels=status_counts.index.tolist(),
            values=status_counts.values.tolist(),
            title="Ticket Status Pipeline",
            subtitle="Current operational lifecycle distribution",
            colors=colors_status,
            center_text=f"{total_issues:,}<br><span style='font-size:10px;color:#64748B;'>TOTAL</span>"
        )
        st.plotly_chart(fig_stat, use_container_width=True)

    with col_d2:
        sla_labels = ['Within TAT', 'Breached TAT', 'Active Open']
        sla_values = [closed_within, closed_without, total_open]
        colors_sla = ['#10B981', '#EF4444', '#F59E0B']
        fig_sla = plot_donut_chart(
            labels=sla_labels,
            values=sla_values,
            title="SLA Compliance Distribution",
            subtitle="Breakdown of target turnaround compliance",
            colors=colors_sla,
            center_text=f"{cm_tat_eff:.2f}%<br><span style='font-size:10px;color:#64748B;'>SLA RATE</span>"
        )
        st.plotly_chart(fig_sla, use_container_width=True)

    with col_d3:
        sev_counts = filtered_df['severity'].value_counts()
        sev_colors = {'Critical': '#DC2626', 'Major': '#F59E0B', 'Minor': '#3B82F6'}
        colors_sev = [sev_colors.get(s, '#64748B') for s in sev_counts.index]
        fig_sev = plot_donut_chart(
            labels=sev_counts.index.tolist(),
            values=sev_counts.values.tolist(),
            title="Severity Matrix",
            subtitle="Fault distribution by incident criticality",
            colors=colors_sev,
            center_text=f"{len(sev_counts)}<br><span style='font-size:10px;color:#64748B;'>LEVELS</span>"
        )
        st.plotly_chart(fig_sev, use_container_width=True)

    # ── Charts Section 3: Root Cause Pareto & OEM Reliability ────────────────
    st.markdown('<div class="section-header">⚠️ <span class="section-header-accent">Failure Root Causes</span> & OEM Reliability</div>', unsafe_allow_html=True)
    col_p1, col_p2 = st.columns([6, 6])

    with col_p1:
        fig_pareto = plot_pareto_root_causes(filtered_df, top_n=8)
        if fig_pareto:
            st.plotly_chart(fig_pareto, use_container_width=True)
        else:
            st.info("No category data available.")

    with col_p2:
        fig_oem = plot_charger_make_analysis(filtered_df)
        if fig_oem:
            st.plotly_chart(fig_oem, use_container_width=True)
        else:
            st.info("No charger OEM make data available.")

    # ── Charts Section 4: Station Hotspots ────────────────────────────────────
    st.markdown('<div class="section-header">📍 <span class="section-header-accent">Station Hotspots</span> Analytics</div>', unsafe_allow_html=True)
    fig_hotspots = plot_station_hotspots(filtered_df, top_n=10)
    if fig_hotspots:
        st.plotly_chart(fig_hotspots, use_container_width=True)
    else:
        st.info("No station data available.")


# ─── View 2: Repetitive Faults Analytics ──────────────────────────────────────

def render_repetitive_faults_view(filtered_df):
    """Render comprehensive repetitive faults analytics (Top 10 Overall & Top 20 by Station)."""
    st.markdown('<div class="section-header">🔁 <span class="section-header-accent">Repetitive Faults</span> &amp; Root Cause Hotspots</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.info("No issue records available for the active filters.")
        return

    top10_overall, top20_station = compute_top_repetitive_faults(filtered_df, top_overall=10, top_station=20)

    # ── KPI Cards for Repetitive Faults ──────────────────────────────────────
    total_incidents = len(filtered_df)
    rep_faults_pool = filtered_df[filtered_df.duplicated(subset=['issueType', 'issueSubType'], keep=False)]
    total_rep_count = len(rep_faults_pool)
    top_rep_cause = top10_overall['issueSubType'].iloc[0] if not top10_overall.empty else 'N/A'
    top_rep_cause_count = top10_overall['total_count'].iloc[0] if not top10_overall.empty else 0
    top_station_hotspot = top20_station['stationName'].iloc[0] if not top20_station.empty else 'N/A'
    top_station_count = top20_station['fault_count'].iloc[0] if not top20_station.empty else 0
    overall_rep_sla = (top10_overall['within_tat'].sum() / top10_overall['total_count'].sum() * 100) if not top10_overall.empty and top10_overall['total_count'].sum() > 0 else 0.0

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f"""
            <div class="metric-card darkred">
                <div class="metric-header">
                    <span class="metric-label">Repetitive Incidents</span>
                    <span class="metric-icon">🔁</span>
                </div>
                <div class="metric-val">{total_rep_count:,}</div>
                <div class="metric-sub">{(total_rep_count/total_incidents*100):.2f}% of Total Network Faults</div>
            </div>
        """, unsafe_allow_html=True)

    with k2:
        st.markdown(f"""
            <div class="metric-card red">
                <div class="metric-header">
                    <span class="metric-label">#1 Failure Mode</span>
                    <span class="metric-icon">⚠️</span>
                </div>
                <div class="metric-val" style="font-size: 1.35rem; line-height: 1.3;">{top_rep_cause}</div>
                <div class="metric-sub"><b style="color:#DC2626;">{top_rep_cause_count:,}</b> Total Occurrences</div>
            </div>
        """, unsafe_allow_html=True)

    with k3:
        st.markdown(f"""
            <div class="metric-card amber">
                <div class="metric-header">
                    <span class="metric-label">#1 Station Hotspot</span>
                    <span class="metric-icon">📍</span>
                </div>
                <div class="metric-val" style="font-size: 1.25rem; line-height: 1.3;">{str(top_station_hotspot)[:26]}...</div>
                <div class="metric-sub"><b style="color:#D97706;">{top_station_count:,}</b> Repeated Incidents</div>
            </div>
        """, unsafe_allow_html=True)

    with k4:
        st.markdown(f"""
            <div class="metric-card blue">
                <div class="metric-header">
                    <span class="metric-label">Top Faults SLA %</span>
                    <span class="metric-icon">⚡</span>
                </div>
                <div class="metric-val">{overall_rep_sla:.2f}%</div>
                <div class="metric-sub">SLA Adherence on Top Faults</div>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Tabs for Repetitive Views ────────────────────────────────────────────
    tab_issue_types, tab_station, tab_drilldown = st.tabs([
        "⚡ Top 10 Issue Types & Top 5 Sub-Types Analytics",
        "📍 Repetitive Faults by Station Location",
        "🔍 Incident Ticket Drilldown"
    ])

    with tab_issue_types:
        st.markdown('<div class="section-header">⚡ <span class="section-header-accent">Top 10 Overall Repetitive Issue Types</span> &amp; Top 5 Sub-Types Explorer</div>', unsafe_allow_html=True)
        st.caption("Click on any Issue Type below or use the dropdown selector to view its Top 5 Sub-Types and affected station/charger records:")

        # Aggregate Top 10 Overall Issue Types
        type_counts = filtered_df.groupby('issueType').size().reset_index(name='Total_Faults').sort_values('Total_Faults', ascending=False)
        top10_types_df = type_counts.head(10).copy()
        tot_all_faults = len(filtered_df)
        top10_types_df['Share %'] = (top10_types_df['Total_Faults'] / tot_all_faults * 100).round(1)
        top10_types_df.rename(columns={'issueType': 'Issue Type'}, inplace=True)

        # Top Overview Row (Bar Chart + Summary Table)
        col_t1, col_t2 = st.columns([6, 6])
        with col_t1:
            fig_types = plot_top_issue_types_barchart(top10_types_df)
            if fig_types:
                st.plotly_chart(fig_types, use_container_width=True)
            else:
                st.info("No issue type data available.")
        with col_t2:
            st.write("##### Top 10 Issue Types Summary Table:")
            st.dataframe(
                top10_types_df.style.format({
                    'Total_Faults': '{:,}',
                    'Share %': '{:.1f}%'
                }),
                use_container_width=True,
                height=340
            )

        st.markdown("<hr style='border-top: 1.5px solid #F1F5F9; margin: 1.5rem 0 1.0rem 0;'>", unsafe_allow_html=True)

        # Expanders View for All Top 10 Issue Types (Matches user screenshot)
        st.markdown("##### 📂 Expanders View for All Top 10 Issue Types")
        for idx, row in top10_types_df.iterrows():
            cat_name = row['Issue Type']
            cat_tot = row['Total_Faults']

            c_slice = filtered_df[filtered_df['issueType'] == cat_name]
            sub_agg = c_slice.groupby('issueSubType').size().reset_index(name='Subtype_Count').sort_values('Subtype_Count', ascending=False)
            t5_sub = sub_agg.head(5).copy()
            t5_sub['Sub-Type Share %'] = (t5_sub['Subtype_Count'] / cat_tot * 100).round(1)
            t5_sub.rename(columns={'issueSubType': 'Issue Sub-Type'}, inplace=True)

            with st.expander(f"⚡ Issue Type: {cat_name} — Total {cat_tot:,} Occurrences ({len(t5_sub)} Top Sub-Types)", expanded=False):
                c1, c2 = st.columns([6, 6])
                with c1:
                    fig_sub_exp = plot_top_subtypes_barchart(t5_sub, cat_name)
                    if fig_sub_exp:
                        st.plotly_chart(fig_sub_exp, use_container_width=True, key=f"chart_exp_cat_{idx}")
                with c2:
                    st.write(f"##### 📌 Top 5 Sub-Types Table for {cat_name}:")
                    st.dataframe(
                        t5_sub.style.format({
                            'Subtype_Count': '{:,}',
                            'Sub-Type Share %': '{:.1f}%'
                        }),
                        use_container_width=True,
                        height=260
                    )

                st.write(f"##### 📝 Underlying Ticket Records for Issue Type '{cat_name}':")
                st.dataframe(c_slice[[c for c in ['ocppId', 'stationId', 'stationName', 'issueSubType', 'status', 'tatCompliance', 'zme', 'zone'] if c in c_slice.columns]], use_container_width=True, height=240)

    with tab_station:
        if not top20_station.empty:
            # Expanders View for All Top 20 Stations
            st.markdown("##### 📂 Expanders View for Top 20 Station Locations")
            for idx, row in top20_station.iterrows():
                st_name = row['stationName']
                st_tot = row['fault_count']
                st_zone = row['zone']
                st_zme = row['zme']

                st_slice = filtered_df[filtered_df['stationName'] == st_name]
                sub_agg = st_slice.groupby('issueSubType').size().reset_index(name='Subtype_Count').sort_values('Subtype_Count', ascending=False)
                t5_sub = sub_agg.head(5).copy()
                t5_sub['Sub-Type Share %'] = (t5_sub['Subtype_Count'] / st_tot * 100).round(1)
                t5_sub.rename(columns={'issueSubType': 'Issue Sub-Type'}, inplace=True)

                with st.expander(f"📍 Station: {st_name} — Total {st_tot:,} Occurrences ({st_zone} | ZME: {st_zme})", expanded=False):
                    c1, c2 = st.columns([6, 6])
                    with c1:
                        fig_sub_st = plot_top_subtypes_barchart(t5_sub, st_name)
                        if fig_sub_st:
                            st.plotly_chart(fig_sub_st, use_container_width=True, key=f"chart_exp_st_{idx}")
                    with c2:
                        st.write(f"##### 📌 Top 5 Sub-Types Table for {st_name}:")
                        st.dataframe(
                            t5_sub.style.format({
                                'Subtype_Count': '{:,}',
                                'Sub-Type Share %': '{:.1f}%'
                            }),
                            use_container_width=True,
                            height=260
                        )

                    st.write(f"##### 📝 Underlying Ticket Records for Station '{st_name}':")
                    st.dataframe(st_slice[[c for c in ['ocppId', 'stationId', 'issueType', 'issueSubType', 'status', 'tatCompliance', 'zme', 'zone'] if c in st_slice.columns]], use_container_width=True, height=240)
        else:
            st.info("No station repetitive fault data available.")

    with tab_drilldown:
        st.write("##### 🔍 Deep-Dive: Raw Incident Logs by Repetitive Fault Mode")
        options_list = top10_overall['fault_label'].tolist() if not top10_overall.empty else []
        if options_list:
            selected_mode = st.selectbox("Select Fault Mode to Inspect Individual Tickets:", options_list, index=0, key="drilldown_fault_mode")
            cat_selected, sub_selected = selected_mode.split(' — ', 1)
            drill_df = filtered_df[(filtered_df['issueType'] == cat_selected) & (filtered_df['issueSubType'] == sub_selected)].copy()

            st.write(f"Displaying **{len(drill_df):,}** individual tickets logged for **{selected_mode}**:")
            disp_drill_cols = [
                'issueId', 'ocppId', 'stationName', 'zone', 'zme', 'severity', 'status',
                'issueDate', 'resolutionDate', 'tatCompliance', 'ageOfIssue', 'description', 'correctiveAction'
            ]
            cols = [c for c in disp_drill_cols if c in drill_df.columns]
            st.dataframe(drill_df[cols], use_container_width=True, height=360)

            csv_drill = drill_df[cols].to_csv(index=False).encode('utf-8')
            st.download_button(
                f"📥 Download Tickets for '{selected_mode}' CSV",
                data=csv_drill,
                file_name=f"Tickets_{cat_selected}_{sub_selected}.csv",
                mime="text/csv",
                key="dl_drilldown_csv"
            )


# ─── View 3: Critical & SLA Breached Triage ───────────────────────────────────

def render_critical_and_breached_view(filtered_df):
    """Focused triage view for high-priority, overdue, and SLA-breached tickets."""
    st.markdown('<div class="section-header">🚨 <span class="section-header-accent">Critical &amp; SLA Breached</span> Incident Triage</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.info("No records matching current filters.")
        return

    crit_df = filtered_df[filtered_df['severity'] == 'Critical']
    overdue_df = filtered_df[filtered_df['_Is_Overdue_']]
    breached_closed_df = filtered_df[filtered_df['_Is_Closed_Without_']]
    all_sla_risk = filtered_df[filtered_df['_Is_Overdue_'] | filtered_df['_Is_Closed_Without_'] | (filtered_df['severity'] == 'Critical')]

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""
            <div class="metric-card red">
                <div class="metric-header">
                    <span class="metric-label">Active Overdue</span>
                    <span class="metric-icon">⏰</span>
                </div>
                <div class="metric-val">{len(overdue_df):,}</div>
                <div class="metric-sub">Open Beyond Target TAT</div>
            </div>
        """, unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
            <div class="metric-card darkred">
                <div class="metric-header">
                    <span class="metric-label">Critical Faults</span>
                    <span class="metric-icon">🔥</span>
                </div>
                <div class="metric-val">{len(crit_df):,}</div>
                <div class="metric-sub">High Priority Severity</div>
            </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
            <div class="metric-card amber">
                <div class="metric-header">
                    <span class="metric-label">Closed Breached</span>
                    <span class="metric-icon">⚠️</span>
                </div>
                <div class="metric-val">{len(breached_closed_df):,}</div>
                <div class="metric-sub">Closed Past SLA Window</div>
            </div>
        """, unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
            <div class="metric-card blue">
                <div class="metric-header">
                    <span class="metric-label">Total SLA Attention</span>
                    <span class="metric-icon">🛡️</span>
                </div>
                <div class="metric-val">{len(all_sla_risk):,}</div>
                <div class="metric-sub">Critical / Breached Pool</div>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    tab_overdue, tab_crit, tab_breached = st.tabs([
        f"⏰ Active Overdue ({len(overdue_df)})",
        f"🔥 Critical Faults ({len(crit_df)})",
        f"⚠️ Closed Breached ({len(breached_closed_df)})"
    ])

    disp_cols = [
        'issueId', 'ocppId', 'stationName', 'zone', 'zme', 'severity', 'status',
        'issueType', 'issueSubType', 'issueDate', 'tdoc', 'ageOfIssue', 'tatDays', 'description'
    ]

    with tab_overdue:
        if not overdue_df.empty:
            cols = [c for c in disp_cols if c in overdue_df.columns]
            st.dataframe(
                overdue_df[cols].rename(columns={
                    'issueId': 'Issue ID', 'ocppId': 'Charger ID', 'stationName': 'Station Name',
                    'zone': 'Zone', 'zme': 'ZME', 'severity': 'Severity', 'status': 'Status',
                    'issueType': 'Category', 'issueSubType': 'Sub-Type', 'issueDate': 'Issue Date',
                    'tdoc': 'Target Date (TDOC)', 'ageOfIssue': 'Age (Days)', 'tatDays': 'Allowed TAT',
                    'description': 'Description'
                }),
                use_container_width=True,
                height=360
            )
            csv_overdue = overdue_df[cols].to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Overdue Tickets CSV", data=csv_overdue, file_name="Overdue_Issues.csv", mime="text/csv", key="dl_overdue_csv")
        else:
            st.success("🎉 No active overdue tickets! All open issues are within target SLA.")

    with tab_crit:
        if not crit_df.empty:
            cols = [c for c in disp_cols if c in crit_df.columns]
            st.dataframe(
                crit_df[cols].rename(columns={
                    'issueId': 'Issue ID', 'ocppId': 'Charger ID', 'stationName': 'Station Name',
                    'zone': 'Zone', 'zme': 'ZME', 'severity': 'Severity', 'status': 'Status',
                    'issueType': 'Category', 'issueSubType': 'Sub-Type', 'issueDate': 'Issue Date',
                    'tdoc': 'Target Date (TDOC)', 'ageOfIssue': 'Age (Days)', 'tatDays': 'Allowed TAT',
                    'description': 'Description'
                }),
                use_container_width=True,
                height=360
            )
            csv_crit = crit_df[cols].to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Critical Issues CSV", data=csv_crit, file_name="Critical_Issues.csv", mime="text/csv", key="dl_crit_csv")
        else:
            st.info("No critical severity issues recorded in the current selection.")

    with tab_breached:
        if not breached_closed_df.empty:
            cols = [c for c in disp_cols if c in breached_closed_df.columns]
            st.dataframe(
                breached_closed_df[cols].rename(columns={
                    'issueId': 'Issue ID', 'ocppId': 'Charger ID', 'stationName': 'Station Name',
                    'zone': 'Zone', 'zme': 'ZME', 'severity': 'Severity', 'status': 'Status',
                    'issueType': 'Category', 'issueSubType': 'Sub-Type', 'issueDate': 'Issue Date',
                    'tdoc': 'Target Date (TDOC)', 'ageOfIssue': 'Age (Days)', 'tatDays': 'Allowed TAT',
                    'description': 'Description'
                }),
                use_container_width=True,
                height=360
            )
            csv_breached = breached_closed_df[cols].to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Closed Breached CSV", data=csv_breached, file_name="Closed_Breached_Issues.csv", mime="text/csv", key="dl_breached_csv")
        else:
            st.success("🎉 No SLA-breached closures in the current selection.")


# ─── View 4: ZME & Zone Governance Scorecard ──────────────────────────────────

def render_governance_scorecard(filtered_df):
    """Detailed scorecards and rankings for ZMEs and Zones."""
    st.markdown('<div class="section-header">👷 <span class="section-header-accent">Field Engineering</span> Governance &amp; Scorecards</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.info("No data available for governance scorecard.")
        return

    tab_zme, tab_zone = st.tabs(["👷 ZME Scorecard Leaderboard", "🏢 Zone Regional Governance"])

    with tab_zme:
        zme_table = compute_zme_issue_table(filtered_df)
        if not zme_table.empty:
            st.write(f"Showing performance metrics for **{len(zme_table)}** Zone Maintenance Engineers:")
            zme_disp = zme_table.rename(columns={
                'zme': 'ZME Engineer',
                'zone': 'Zone',
                'total': 'Registered',
                'open': 'Open',
                'closed': 'Closed',
                'within': 'Within TAT',
                'outside': 'Breached TAT',
                'overdue': 'Overdue',
                'critical': 'Critical',
                'cm_efficiency': 'CM Eff %',
                'tat_efficiency': 'SLA %'
            })
            st.dataframe(
                zme_disp.style.format({
                    'Registered': '{:,}',
                    'Open': '{:,}',
                    'Closed': '{:,}',
                    'Within TAT': '{:,}',
                    'Breached TAT': '{:,}',
                    'Overdue': '{:,}',
                    'Critical': '{:,}',
                    'CM Eff %': '{:.2f}%',
                    'SLA %': '{:.2f}%'
                }),
                use_container_width=True,
                height=420
            )

            csv_zme = zme_table.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download ZME Scorecard CSV", data=csv_zme, file_name="ZME_Performance_Scorecard.csv", mime="text/csv", key="dl_zme_csv")
        else:
            st.info("No ZME records available.")

    with tab_zone:
        zone_table = compute_zone_issue_table(filtered_df)
        if not zone_table.empty:
            st.write(f"Showing regional performance across **{len(zone_table)}** Zones:")
            zone_disp = zone_table.rename(columns={
                'zone': 'Zone / Region',
                'total': 'Total Registered',
                'open': 'Open',
                'closed': 'Closed',
                'within': 'Within TAT',
                'outside': 'Breached TAT',
                'overdue': 'Overdue',
                'zme_count': 'ZME Count',
                'station_count': 'Impacted Sites',
                'cm_efficiency': 'CM Eff %',
                'tat_efficiency': 'SLA %'
            })
            st.dataframe(
                zone_disp.style.format({
                    'Total Registered': '{:,}',
                    'Open': '{:,}',
                    'Closed': '{:,}',
                    'Within TAT': '{:,}',
                    'Breached TAT': '{:,}',
                    'Overdue': '{:,}',
                    'ZME Count': '{:,}',
                    'Impacted Sites': '{:,}',
                    'CM Eff %': '{:.2f}%',
                    'SLA %': '{:.2f}%'
                }),
                use_container_width=True,
                height=340
            )

            csv_zone = zone_table.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Download Zone Scorecard CSV", data=csv_zone, file_name="Zone_Performance_Summary.csv", mime="text/csv", key="dl_zone_csv")
        else:
            st.info("No Zone data available.")


# ─── View 5: Issue Explorer ───────────────────────────────────────────────────

def render_issue_explorer(filtered_df):
    """Raw issue data grid with custom column selector and filtering."""
    st.markdown('<div class="section-header">📋 <span class="section-header-accent">Issue Tracker</span> Data Explorer</div>', unsafe_allow_html=True)

    if filtered_df.empty:
        st.info("No issue records loaded or matching current filters.")
        return

    st.write(f"Displaying **{len(filtered_df):,}** operational issue records matching active filters:")

    all_cols = [c for c in filtered_df.columns if not c.startswith('_')]
    default_cols = [
        'issueId', 'ocppId', 'stationName', 'zone', 'zme', 'severity', 'status',
        'issueType', 'issueSubType', 'issueDate', 'resolutionDate', 'tatCompliance', 'ageOfIssue', 'chargerMake'
    ]
    selected_cols = st.multiselect(
        "Select columns to display:",
        all_cols,
        default=[c for c in default_cols if c in all_cols],
        key="exp_col_picker"
    )

    disp_cols = selected_cols if selected_cols else all_cols
    st.dataframe(filtered_df[disp_cols], use_container_width=True, height=480)

    csv_data = filtered_df[disp_cols].to_csv(index=False).encode('utf-8')
    st.download_button(
        label="📥 Download Filtered Issues CSV",
        data=csv_data,
        file_name="Filtered_Issues_Data.csv",
        mime="text/csv",
        key="dl_explorer_csv"
    )


# ─── View 6: Data Upload & Validation ─────────────────────────────────────────

def render_data_upload():
    """Upload and validate Issue Tracker workbooks."""
    st.markdown('<div class="section-header">📤 <span class="section-header-accent">Issue Tracker</span> Workbook Upload</div>', unsafe_allow_html=True)

    st.markdown("""
        <p style="color: #475569; font-size: 0.98rem; margin-bottom: 1.5rem; line-height: 1.6;">
            Upload your operational <b>Issue Tracker (F-02-MAINT-001)</b> workbook in <code>.xlsx</code> format.
            The system automatically parses all fields including SLA compliance, ZME allocations, resolution timestamps, and root causes.
        </p>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div class="upload-card">
            <div style="font-size: 3.4rem; margin-bottom: 0.6rem;">🗂️</div>
            <h3 style="color: #991B1B !important; margin-bottom: 0.3rem;">F-02 Issue Tracker Upload</h3>
            <p style="color: #64748B; font-size: 0.90rem; max-width: 580px; margin: 0 auto 1.2rem; line-height: 1.5;">
                Expected worksheet: <b>Issue Tracker</b> with columns like <code>Issue Id</code>, <code>OCPP ID</code>,
                <code>Issue Date</code>, <code>Severity</code>, <code>Status</code>, <code>ZME</code>, <code>Zone</code>,
                <code>TAT Compliance</code>, and <code>Charger Make</code>.
            </p>
        </div>
    """, unsafe_allow_html=True)

    issue_file = st.file_uploader(
        "Upload Issue Tracker (.xlsx)",
        type=["xlsx"],
        key="upload_issue_file_tab",
        label_visibility="collapsed"
    )

    if issue_file:
        st.session_state['uploaded_issue_bytes'] = issue_file.getvalue()
        st.session_state['uploaded_file_name'] = issue_file.name
        st.success(f"✅ **{issue_file.name}** uploaded successfully ({issue_file.size // 1024:,} KB). Parsing...")
        st.rerun()

    # Column Mapping Preview
    uploaded_bytes = st.session_state.get('uploaded_issue_bytes')
    if uploaded_bytes:
        st.markdown('<div class="section-header">🔍 <span class="section-header-accent">Column Mapping</span> Validation</div>', unsafe_allow_html=True)
        try:
            wb = openpyxl.load_workbook(BytesIO(uploaded_bytes), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            sheet_target = select_sheet_name(sheet_names, ['Issue Tracker', 'Issue Data'], 'issue')
            ws = wb[sheet_target]
            rows = list(ws.iter_rows(values_only=True, max_row=5))
            wb.close()

            if rows:
                headers = []
                for r in rows:
                    if r and any(c is not None for c in r):
                        headers = [str(c).strip() for c in r if c is not None]
                        break

                expected_cols = [
                    'Issue Id', 'OCPP ID', 'Severity', 'Issue Date', 'Status', 'Issue Type',
                    'Issue Sub-Type', 'Resolution Date', 'Zone', 'ZME', 'Station Name',
                    'TAT Days', 'TDOC', 'TAT Compliance', 'Charger Make'
                ]

                mapping_records = []
                for h in headers:
                    norm_h = norm_header(h)
                    match_found = next((e for e in expected_cols if norm_header(e) in norm_h or norm_h in norm_header(e)), None)
                    mapping_records.append({
                        'Detected Column in File': h,
                        'System Field Match': f"✅ {match_found}" if match_found else "—"
                    })

                map_df = pd.DataFrame(mapping_records)
                st.dataframe(map_df, use_container_width=True, height=280)
        except Exception as e:
            st.warning(f"Could not preview column mapping: {e}")


# ─── Main Application Entry Point ─────────────────────────────────────────────

def main():
    inject_modern_css()
    active_nav = render_header_and_nav()

    # Data Loader strictly from user upload
    uploaded_bytes = st.session_state.get('uploaded_issue_bytes')
    issue_df = load_issue_data(uploaded_bytes)

    if active_nav == "📤 Data Upload & Validation":
        render_data_upload()
        return

    # Onboarding View when no file is uploaded yet
    if issue_df.empty:
        st.markdown("""
            <div style="background: #FFFFFF; border: 1.5px solid #FEE2E2; border-radius: 20px;
                        padding: 3rem 2.5rem; text-align: center; margin-top: 1rem;
                        box-shadow: 0 8px 24px rgba(153,27,27,0.06);">
                <div style="font-size: 3.5rem; margin-bottom: 0.6rem;">⚡</div>
                <h2 style="color: #991B1B !important; font-weight: 900; margin-bottom: 0.5rem; font-size: 1.85rem;">
                    Upload Issue Tracker Workbook
                </h2>
                <p style="color: #475569 !important; font-size: 1.02rem; max-width: 580px; margin: 0 auto 1.5rem; line-height: 1.6;">
                    Upload your <b>F-02-MAINT-001 Issue Tracker</b> spreadsheet (<code>.xlsx</code>) below to generate real-time SLA governance, field engineering scorecards, and failure analytics.
                </p>
            </div>
        """, unsafe_allow_html=True)

        col_u_l, col_u_m, col_u_r = st.columns([2, 8, 2])
        with col_u_m:
            main_uploaded_file = st.file_uploader(
                "Upload Issue Tracker (.xlsx)",
                type=["xlsx"],
                key="main_home_uploader",
                help="Select or drag and drop your F-02-MAINT-001 Issue Tracker Excel file"
            )
            if main_uploaded_file:
                st.session_state['uploaded_issue_bytes'] = main_uploaded_file.getvalue()
                st.session_state['uploaded_file_name'] = main_uploaded_file.name
                st.rerun()

            if os.path.exists('Issue_Tracker.xlsx'):
                st.markdown("<div style='text-align: center; margin-top: 0.8rem; margin-bottom: 0.8rem; color: #64748B; font-weight: 600; font-size: 0.85rem;'>— OR —</div>", unsafe_allow_html=True)
                if st.button("⚡ Load Workspace Sample Dataset (Issue_Tracker.xlsx)", key="btn_load_workspace_sample", use_container_width=True):
                    with open('Issue_Tracker.xlsx', 'rb') as f:
                        st.session_state['uploaded_issue_bytes'] = f.read()
                        st.session_state['uploaded_file_name'] = 'Issue_Tracker.xlsx'
                    st.rerun()
        return

    # Loaded File Status Header
    file_name = st.session_state.get('uploaded_file_name', 'Issue_Tracker.xlsx (Workspace Dataset)')
    col_info, col_reset = st.columns([10, 2])
    with col_info:
        st.markdown(f"""
            <div style="font-size: 0.85rem; color: #475569; font-weight: 600; margin-bottom: 0.5rem; display: flex; align-items: center; gap: 8px;">
                <span>📁 Active File: <b style="color: #0F172A;">{file_name}</b> ({len(issue_df):,} records loaded)</span>
            </div>
        """, unsafe_allow_html=True)
    with col_reset:
        if st.button("🔄 Upload New File", key="btn_change_file", use_container_width=True):
            st.session_state.pop('uploaded_issue_bytes', None)
            st.session_state.pop('uploaded_file_name', None)
            st.rerun()

    # Filter Toolbar
    selected_months, selected_states, selected_zones, selected_zmes, selected_make, search_q = render_filter_toolbar(issue_df)

    # Filtered Dataset
    filtered_df = apply_filters(
        issue_df,
        months=selected_months,
        states=selected_states,
        zones=selected_zones,
        zmes=selected_zmes,
        make=selected_make,
        query=search_q
    )

    # Route Views
    if active_nav == "📊 Executive Analytics":
        render_executive_analytics(filtered_df, issue_df)

    elif active_nav == "🔁 Repetitive Faults":
        render_repetitive_faults_view(filtered_df)

    elif active_nav == "🚨 Critical & SLA Breached":
        render_critical_and_breached_view(filtered_df)

    elif active_nav == "👷 ZME & Zone Governance":
        render_governance_scorecard(filtered_df)

    elif active_nav == "📋 Issue Explorer":
        render_issue_explorer(filtered_df)


if __name__ == '__main__':
    main()
